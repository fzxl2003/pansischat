# coding=utf-8
# This is a sample Python script.
# !/usr/bin/python
# -*- coding: UTF-8 -*-
import os
import sys
import threading
import tkinter as tk
from tkinter import FLAT, END
from socket import socket,AF_INET,SOCK_DGRAM

from cryptography.fernet import Fernet

from openpyxl import load_workbook, Workbook
import time
import win32api
import win32con



udps=socket(AF_INET,SOCK_DGRAM)   #创建套接字
udps.bind(('',8081))   #本机软件绑定端口
ip=('127.0.0.1',8080)    #服务器端ip
udps.settimeout(3)       #设置等待服务器时间时长
udpsituation1=0          #检测网络状况的变量
logcode=b'w2V8gAUFER-9uq6lKG7AfpYI6AcINEs7Ipm4KyQj2AI='   #登录密码秘钥


window=tk.Tk()
window.overrideredirect(True)
window.title('登录页面')
window.geometry('510x258+400+200')
codeshowvariable=tk.StringVar()    #gui窗口创建
indexbgimage= tk.PhotoImage(file="组件/ui/1/登录背景1.png")
offimage0=tk.PhotoImage(file="组件/ui/1/关闭0.png")
minimage0=tk.PhotoImage(file="组件/ui/1/最小化0.png")
offimage1=tk.PhotoImage(file="组件/ui/1/关闭1.png")
minimage1=tk.PhotoImage(file="组件/ui/1/最小化1.png")
uiid = '1'
uimain = '1'

indexbgi=tk.Label(window,image=indexbgimage)
indexbgi.place(x=-5,y=-5)
global usernicheng
userinput=tk.Entry(window,bg='#555555',relief=FLAT,width=25,text='Pansis账户',fg='#8E8E84')
codeinput=tk.Entry(window,bg='#555555',relief=FLAT,width=25,show='',fg='#8E8E84')
udpsituation=tk.Label(window,text='正在连接服务器',fg='#8E8E84',bg='black')
result=tk.Text(window,width=60,height=30,bg='#2B2B2B',fg='#FFFFFF',bd=0)
windowoff=tk.Label(window,image=offimage0,bg='black')

windowmin=tk.Label(window,image=minimage0,bg='black')
setput=tk.Text(window,width=60,height=5,bg='#535353',bd=0)
window.attributes("-alpha", 0.85)

userinput.place(x=300,y=50)
codeinput.place(x=300,y=85)

udpsituation.place(x=310,y=115)
windowoff.place(x=465,y=2)
windowmin.place(x=435,y=2)





def udpsitiontest():
    global udpsituation1
    ewqxd = 1

    try:
           udps.sendto('0000'.encode('gb2312'), ip)
           receivedate = udps.recvfrom(1024)
    except:
           udpsituation.config(text='连接服务器失败')
           udpsituation1=0
           print('失败')

    else:

           udpsituation.config(text='连接服务器成功')
           udpsituation1 = 1


#def liaotiansend(event):         ##聊天输入框回车事件
   # result = setput.get('1.0','end')
  #  print(result)

def encrypt1(date,cipher_key):     # 进行加密
    date = bytes(date, encoding='utf-8')
    encrypted_text = Fernet(cipher_key).encrypt(date)
    encrypted_text = str(encrypted_text, encoding='utf-8')
    return encrypted_text

def decrypt1(date, cipher_key):  # 进行解密
    if date[:2]=='b\'':
        print('123')
        date=date[2:]
        date=date[:-1]
        print(date)
    date = bytes(date, encoding='utf-8')
    print(date)
    decrypted_text = Fernet(cipher_key).decrypt(date)
    decrypted_text = str(decrypted_text, encoding='utf-8')
    print(decrypted_text)
    return decrypted_text

def codeshow1():            #控制密码框显示
    if codeshowvariable.get()=='0':
        codeinput.config(show='*')
    else:codeinput.config(show='')
    return
def newsend(event):
  if udpsituation1 == 1 or udpsituation1 == 0:
    print('1')
    datetime=time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    date=setput.get('1.0', 'end')
    result.config(state='normal')
    result.insert('end', usernicheng + '(我)   ' + datetime + '\n' + date + '\n')
    result.config(state='disable')
    date1='0011'+nowfriendid+'&'+userid+'&'+datetime+'&'+date
    udps.sendto(date1.encode('gb2312'), ip)
    setput.delete('1.0', 'end')
    for col in ws.iter_cols(min_row=2, min_col=1, max_row=9999, max_col=1):
        for cell in col:
            if cell.value==nowfriendid:
                cellrow=cell.row
                break
    for row in ws.iter_rows(min_row=cellrow, min_col=11, max_row=cellrow, max_col=9999):  # 括号代表遍历第一行到第二行,第二列到第三列
        for cell in row:
            if cell.value==None:
                date=datetime+'&'+userid+date
                date=encrypt1(date,code)
                print('而我却'+date)
                cell.value=date
                wb.save(userid+'.xlsx')
                break
  else:
        result.config(state='normal')
        result.insert('end', '网络连接失败\n')
        result.delete('\n')
  setput.focus_force()
  win32api.keybd_event(38, 0, 0, 0)
  win32api.keybd_event(38, 0, win32con.KEYEVENTF_KEYUP, 0)
  result.see(tk.END)
  result.config(state='disable')
  return
def mousemove(event):
    global friendlistmainy
    global friendlistmain
    global friendlistmainsum
    if event.delta > 1 and friendlistmainsum!=1:
        friendlistmainy=friendlistmainy+10
    if event.delta < 1 and friendlistmainsum!=len(friendline)-7:
        friendlistmainy=friendlistmainy-10
    #print(friendlistmainy)
    if friendlistmainy<= friendlistmainymin-40:
        friendlistmainsum=friendlistmainsum+1
        friendlistmain=friendlistmain+1
        friendlistmainy=friendlistmainy+45

        print('friendlistmain'+str(friendlistmain))
        print('friendlistmainsum'+str(friendlistmainsum))
        if friendlistmain==9:
            friendlistmain=1
        friendlistmoveup()
            #print(friendlistmain)

    if friendlistmainy>=friendlistmainymin+40:
        friendlistmainsum = friendlistmainsum - 1
        friendlistmain=friendlistmain-1
        friendlistmainy=friendlistmainy-45

        if friendlistmain==0:
            friendlistmain=8
        friendlistmovedown()
    friendlistmove()


def friendlistmovedown():
    global friendlistmain
    global friendlistmainsum
    global friendlist8id
    global friendlist1id
    global friendlist2id
    global friendlist3id
    global friendlist4id
    global friendlist5id
    global friendlist6id
    global friendlist7id
    global friendline

    if friendlistmain == 1:
        friendlist1id = friendlistmainsum
        friendlist1.config(text=str(friendline[friendlist1id-1]))
        #print(friendlist1id)
    if friendlistmain == 2:
        friendlist2id = friendlistmainsum
        friendlist2.config(text=str(friendline[friendlist2id-1]))
        #print(friendlist2id)
    if friendlistmain == 3:
        friendlist3id = friendlistmainsum
        friendlist3.config(text=str(friendline[friendlist3id-1]))
        #print(friendlist3id)
    if friendlistmain == 4:
        friendlist4id = friendlistmainsum
        friendlist4.config(text=str(friendline[friendlist4id-1]))
        #print(friendlist4id)
    if friendlistmain == 5:
        friendlist5id = friendlistmainsum
        friendlist5.config(text=str(friendline[friendlist5id-1]))
        #print(friendlist5id)
    if friendlistmain == 6:
        friendlist6id = friendlistmainsum
        friendlist6.config(text=str(friendline[friendlist6id-1]))
        #print(friendlist6id)
    if friendlistmain == 7:
        friendlist7id = friendlistmainsum
        friendlist7.config(text=str(friendline[friendlist7id-1]))
        #print(friendlist7id)
    if friendlistmain == 8:
        friendlist8id = friendlistmainsum
        friendlist8.config(text=str(friendline[friendlist8id-1]))
        #print(friendlist8id)

def friendlistmoveup():
    global friendlistmain
    global friendlistmainsum
    global friendlist8id
    global friendlist1id
    global friendlist2id
    global friendlist3id
    global friendlist4id
    global friendlist5id
    global friendlist6id
    global friendlist7id
    global friendline


    if friendlistmain == 1:
        friendlist8id = friendlistmainsum + 7
        friendlist8.config(text=str(friendline[friendlist8id-1]))
        #print(friendlist8id)
    if friendlistmain == 2:
        friendlist1id = friendlistmainsum + 7
        friendlist1.config(text=str(friendline[friendlist1id-1]))
        #print(friendlist1id)
    if friendlistmain == 3:
            friendlist2id = friendlistmainsum + 7
            friendlist2.config(text=str(friendline[friendlist2id - 1]))
         #   print(friendlist2id)
    if friendlistmain == 4:
        friendlist3id = friendlistmainsum + 7
        friendlist3.config(text=str(friendline[friendlist3id-1]))
        #print(friendlist3id)
    if friendlistmain == 5:
        friendlist4id = friendlistmainsum + 7
        friendlist4.config(text=str(friendline[friendlist4id-1]))
        #print(friendlist4id)
    if friendlistmain == 6:
        friendlist5id = friendlistmainsum + 7
        friendlist5.config(text=str(friendline[friendlist5id-1]))
        #print(friendlist5id)
    if friendlistmain == 7:
        friendlist6id = friendlistmainsum + 7
        friendlist6.config(text=str(friendline[friendlist6id-1]))
        #print(friendlist6id)
    if friendlistmain == 8:
        #print('friendlistmainsum' + str(friendlistmainsum))
        friendlist7id = friendlistmainsum + 7
        friendlist7.config(text=str(friendline[friendlist7id-1]))





def friendlistmove():   #好友列表滚动
    if friendlistmain==1:
        friendlist1.place(x=friendlistmainx, y=friendlistmainy)
        friendlist2.place(x=friendlistmainx, y=friendlistmainy+45*1)
        friendlist3.place(x=friendlistmainx, y=friendlistmainy + 45*2)
        friendlist4.place(x=friendlistmainx, y=friendlistmainy + 45*3)
        friendlist5.place(x=friendlistmainx, y=friendlistmainy + 45*4)
        friendlist6.place(x=friendlistmainx, y=friendlistmainy + 45*5)
        friendlist7.place(x=friendlistmainx, y=friendlistmainy + 45*6)
        friendlist8.place(x=friendlistmainx, y=friendlistmainy + 45*7)

    if friendlistmain==2:
        friendlist2.place(x=friendlistmainx, y=friendlistmainy)
        friendlist3.place(x=friendlistmainx, y=friendlistmainy+45*1)
        friendlist4.place(x=friendlistmainx, y=friendlistmainy + 45*2)
        friendlist5.place(x=friendlistmainx, y=friendlistmainy + 45*3)
        friendlist6.place(x=friendlistmainx, y=friendlistmainy + 45*4)
        friendlist7.place(x=friendlistmainx, y=friendlistmainy + 45*5)
        friendlist8.place(x=friendlistmainx, y=friendlistmainy + 45*6)
        friendlist1.place(x=friendlistmainx, y=friendlistmainy + 45*7)
    if friendlistmain==3:
        friendlist3.place(x=friendlistmainx, y=friendlistmainy)
        friendlist4.place(x=friendlistmainx, y=friendlistmainy+45*1)
        friendlist5.place(x=friendlistmainx, y=friendlistmainy + 45*2)
        friendlist6.place(x=friendlistmainx, y=friendlistmainy + 45*3)
        friendlist7.place(x=friendlistmainx, y=friendlistmainy + 45*4)
        friendlist8.place(x=friendlistmainx, y=friendlistmainy + 45*5)
        friendlist1.place(x=friendlistmainx, y=friendlistmainy + 45*6)
        friendlist2.place(x=friendlistmainx, y=friendlistmainy + 45*7)
    if friendlistmain==4:
        friendlist4.place(x=friendlistmainx, y=friendlistmainy)
        friendlist5.place(x=friendlistmainx, y=friendlistmainy+45*1)
        friendlist6.place(x=friendlistmainx, y=friendlistmainy + 45*2)
        friendlist7.place(x=friendlistmainx, y=friendlistmainy + 45*3)
        friendlist8.place(x=friendlistmainx, y=friendlistmainy + 45*4)
        friendlist1.place(x=friendlistmainx, y=friendlistmainy + 45*5)
        friendlist2.place(x=friendlistmainx, y=friendlistmainy + 45*6)
        friendlist3.place(x=friendlistmainx, y=friendlistmainy + 45*7)
    if friendlistmain==5:
        friendlist5.place(x=friendlistmainx, y=friendlistmainy)
        friendlist6.place(x=friendlistmainx, y=friendlistmainy+45*1)
        friendlist7.place(x=friendlistmainx, y=friendlistmainy + 45*2)
        friendlist8.place(x=friendlistmainx, y=friendlistmainy + 45*3)
        friendlist1.place(x=friendlistmainx, y=friendlistmainy + 45*4)
        friendlist2.place(x=friendlistmainx, y=friendlistmainy + 45*5)
        friendlist3.place(x=friendlistmainx, y=friendlistmainy + 45*6)
        friendlist4.place(x=friendlistmainx, y=friendlistmainy + 45*7)
    if friendlistmain==6:
        friendlist6.place(x=friendlistmainx, y=friendlistmainy)
        friendlist7.place(x=friendlistmainx, y=friendlistmainy+45*1)
        friendlist8.place(x=friendlistmainx, y=friendlistmainy + 45*2)
        friendlist1.place(x=friendlistmainx, y=friendlistmainy + 45*3)
        friendlist2.place(x=friendlistmainx, y=friendlistmainy + 45*4)
        friendlist3.place(x=friendlistmainx, y=friendlistmainy + 45*5)
        friendlist4.place(x=friendlistmainx, y=friendlistmainy + 45*6)
        friendlist5.place(x=friendlistmainx, y=friendlistmainy + 45*7)
    if friendlistmain==7:
        friendlist7.place(x=friendlistmainx, y=friendlistmainy)
        friendlist8.place(x=friendlistmainx, y=friendlistmainy+45*1)
        friendlist1.place(x=friendlistmainx, y=friendlistmainy + 45*2)
        friendlist2.place(x=friendlistmainx, y=friendlistmainy + 45*3)
        friendlist3.place(x=friendlistmainx, y=friendlistmainy + 45*4)
        friendlist4.place(x=friendlistmainx, y=friendlistmainy + 45*5)
        friendlist5.place(x=friendlistmainx, y=friendlistmainy + 45*6)
        friendlist6.place(x=friendlistmainx, y=friendlistmainy + 45*7)
    if friendlistmain==8:
        friendlist8.place(x=friendlistmainx, y=friendlistmainy)
        friendlist1.place(x=friendlistmainx, y=friendlistmainy+45*1)
        friendlist2.place(x=friendlistmainx, y=friendlistmainy + 45*2)
        friendlist3.place(x=friendlistmainx, y=friendlistmainy + 45*3)
        friendlist4.place(x=friendlistmainx, y=friendlistmainy + 45*4)
        friendlist5.place(x=friendlistmainx, y=friendlistmainy + 45*5)
        friendlist6.place(x=friendlistmainx, y=friendlistmainy + 45*6)
        friendlist7.place(x=friendlistmainx, y=friendlistmainy + 45*7)





def friendlist1click(event):
    global nowfriendid
    print('1')
    nowfriendid=friendline[friendlist1id-1]
    friendlistclick()


def friendlist2click(event):
    global nowfriendid
    nowfriendid = friendline[friendlist2id - 1]
    friendlistclick()
def friendlist3click(event):
    global nowfriendid
    nowfriendid = friendline[friendlist3id - 1]
    friendlistclick()
def friendlist4click(event):
    global nowfriendid
    nowfriendid = friendline[friendlist4id - 1]
    friendlistclick()
def friendlist5click(event):
    global nowfriendid
    nowfriendid = friendline[friendlist5id - 1]
    friendlistclick()
def friendlist6click(event):
    global nowfriendid
    nowfriendid = friendline[friendlist6id - 1]
    friendlistclick()
def friendlist7click(event):
    global nowfriendid
    nowfriendid = friendline[friendlist7id - 1]
    friendlistclick()
def friendlist8click(event):
    global nowfriendid
    nowfriendid = friendline[friendlist8id - 1]
    friendlistclick()

def friendlistclick():

    result.config(state='disable')
    global usernicheng
    global friendnicheng
    global nowfriendrow

    result.place(x=350, y=45)
    setput.place(x=350, y=450)
    for col in ws.iter_cols(min_row=2, min_col=1, max_col=1,max_row=999):
     for cell in col:
        if cell.value==nowfriendid:
            nowfriendrow=cell.row
            usernicheng=ws['B1'].value
            break
    result.config(state='normal')
    result.delete('1.0', 'end')
    udp('0013'+nowfriendid)
def friendlistclick2():
    global usernicheng
    global friendnicheng
    global nowfriendrow
    for row in ws.iter_rows(min_row=nowfriendrow, min_col=11, max_row=nowfriendrow,max_col=999):  # 括号代表遍历第一行到第二行,第二列到第三列
        for cell in row:
            if cell.value!=None:

                date=decrypt1(cell.value,code)
                print(date)
                i=0
                while i<=len(date)-1:
                    if date[i]=='&':
                        print(date)
                        datetime=date[:i]
                        dateid=date[i+1:i+6]
                        date=date[i+6:]
                        if dateid==userid:
                            dateid=usernicheng+'(我)'

                        else:
                            dateid = dateid +'  '+str(friendnicheng)[2:]+ '(对方)'
                        print(date)
                        result.insert('end',dateid+'   '+datetime+'\n'+date+'\n')
                        break
                    i=i+1
    setput.focus_force()
    win32api.keybd_event(38, 0, 0, 0)
    win32api.keybd_event(38, 0, win32con.KEYEVENTF_KEYUP, 0)

    result.see(tk.END)
    result.config(state='disable')
    setput.bind('<Return>',newsend)








def shezhi(event):
    t1 = threading.Thread(target=shezhiwindow)
    t1.setDaemon(True)
    t1.start()


def shezhiwindow():
    global jinggao
    global window2
    global slabel1
    global slabel2
    global slabel3
    global slabel4
    global slabel5
    global slabel6
    global slabel7
    global soff
    global simagemax
    global scodeentry
    global snichengentry
    global sleft
    global sright
    global sbaocun
    global simagemin
    jinggao=0

    window2 = tk.Toplevel()
    window2.overrideredirect(True)
    window2.title('设置')
    window2.geometry('300x530+400+200')
    simagemax = tk.Label(window2)
    slabel1 = tk.Label(window2, text='Pansis账户ID')
    slabel2 = tk.Label(window2, text=userid)
    slabel3 = tk.Label(window2, text='账户删除')
    slabel4 = tk.Label(window2, text='密码')
    slabel5 = tk.Label(window2, text='昵称')
    slabel6 = tk.Label(window2, text='主题')
    slabel7 = tk.Label(window2, text='注意所有聊天记录将全部删除且无法找回',fg='red')
    scodeentry = tk.Entry(window2, width=15, bd=0)
    snichengentry = tk.Entry(window2, width=15, bd=0)
    snichengentry.delete(0, END)
    snichengentry.insert(END,ws["B1"].value)
    scodeentry.delete(0, END)
    # image1=tk.PhotoImage(file="组件/ui/1/展示背景1.png")
    # image2=tk.PhotoImage(file="组件/ui/1/展示背景2.png")
    # image3=tk.PhotoImage(file="组件/ui/2/展示背景3.png")
    # image4=tk.PhotoImage(file="组件/ui/2/展示背景4.png")
    sleft = tk.Label(window2)
    sright = tk.Label(window2)
    soff = tk.Label(window2)
    simagemin = tk.Label(window2)

    sbaocun = tk.Button(window2, bg='#298CF4', relief=FLAT, width=25, command=baocun1, text='保存', fg='white', bd=0,activebackground='#298CF4')
    simagemax.place(x=-2, y=-2)
    slabel1.place(x=5, y=50)
    slabel2.place(x=100, y=50)
    slabel3.place(x=160, y=50)
    slabel4.place(x=5, y=100)
    slabel5.place(x=5, y=150)
    slabel6.place(x=5, y=200)
    soff.place(x=1, y=3)
    sleft.place(x=5, y=240)
    sright.place(x=245, y=240)
    simagemin.place(x=35, y=230)

    sbaocun.place(x=30, y=450)
    scodeentry.place(x=100, y=100)
    snichengentry.place(x=100, y=150)

    # offimage11 = tk.PhotoImage(file="组件/ui/1/关闭1.png")
    # offimage10 = tk.PhotoImage(file="组件/ui/1/关闭0.png")
    # offimage11 = tk.PhotoImage(file="组件/ui/1/关闭1.png")
    soff.bind("<Button-1>", swindowoff1)

    slabel3.bind("<Button-1>", suserdelete)
    soff.bind("<Enter>", swindowoff2)
    soff.bind("<Leave>", swindowoff3)
    sleft.bind("<Button-1>", swindowleft1)
    sleft.bind("<Enter>", swindowleft2)
    sleft.bind("<Leave>", swindowleft3)
    sright.bind("<Button-1>", swindowright1)
    sright.bind("<Enter>", swindowright2)
    sright.bind("<Leave>", swindowright3)
    sui()



def sui():

    global soffimage0
    global soffimage1
    global srightimage0
    global srightimage1
    global sleftimage0
    global sleftimage1
    global simagemaximage
    global simageminimage


    if int(uiid) <= 2:
        uimain = '1'
    else:
        uimain = '2'
    print('uiid:'+uiid)
    print('uiid:'+uimain)
    simagemaximage = tk.PhotoImage(file="组件/ui/" + uimain + "/注册背景" + uiid + ".png")
    simageminimage = tk.PhotoImage(file="组件/ui/" + uimain + "/展示背景" + uiid + ".png")
    print("组件/ui/" + uimain + "/展示背景" + uiid + ".png")
    print("组件/ui/" + uimain + "/注册背景" + uiid + ".png")
    soffimage0 = tk.PhotoImage(file="组件/ui/" + uimain + "/返回0.png")
    soffimage1 = tk.PhotoImage(file="组件/ui/" + uimain + "/返回1.png")
    sleftimage0 = tk.PhotoImage(file="组件/ui/" + uimain + "/左翻页0.png")
    sleftimage1 = tk.PhotoImage(file="组件/ui/" + uimain + "/左翻页1.png")
    srightimage0 = tk.PhotoImage(file="组件/ui/" + uimain + "/右翻页0.png")
    srightimage1 = tk.PhotoImage(file="组件/ui/" + uimain + "/右翻页1.png")
    simagemax.config(image=simagemaximage)
    simagemin.config(image=simageminimage)
    sright.config(image=srightimage0)
    sleft.config(image=sleftimage0)
    if uimain == '1':
        soff.config(image=soffimage0, bg='black')
        scodeentry.config(bg='#555555')
        snichengentry.config(bg='#555555')
        slabel1.config(bg='black', fg='#8E8E84')
        slabel2.config(bg='black', fg='#8E8E84')
        slabel3.config(bg='black', fg='#8E8E84')
        slabel4.config(bg='black', fg='#8E8E84')
        slabel5.config(bg='black', fg='#8E8E84')
        slabel6.config(bg='black', fg='#8E8E84')
        sleft.config(bg='black')
        sright.config(bg='black')
        simagemin.config(bg='black')
        simagemax.config(bg='black')
        window2.config(bg='black')
    if uimain == '2':
            soff.config(image=soffimage0, bg='white')
            scodeentry.config(bg='#F0F0F0')
            snichengentry.config(bg='#F0F0F0')
            slabel1.config(bg='white', fg='#8E8E84')
            slabel2.config(bg='white', fg='#8E8E84')
            slabel3.config(bg='white', fg='#8E8E84')
            slabel4.config(bg='white', fg='#8E8E84')
            slabel5.config(bg='white', fg='#8E8E84')
            slabel6.config(bg='white', fg='#8E8E84')
            sleft.config(bg='white')
            sright.config(bg='white')
            simagemin.config(bg='white')
            simagemax.config(bg='white')
            window2.config(bg='white')






def swindowoff2(event):
    soff.config(image=soffimage1)


def swindowoff3(event):
    soff.config(image=soffimage0)


def swindowoff1(event):  ###
    print('1')
    window2.destroy()


def swindowright2(event):
    sright.config(image=srightimage1)


def swindowright3(event):
    sright.config(image=srightimage0)


def swindowright1(event):  ###
    global uiid
    suiid = int(uiid)
    if suiid < 4:
        suiid = suiid + 1
        uiid = str(suiid)
        sui()
        userui(uiid)


def swindowleft2(event):
    sleft.config(image=sleftimage1)


def swindowleft3(event):
    sleft.config(image=sleftimage0)


def swindowleft1(event):  ###
    global uiid
    suiid=int(uiid)
    if suiid>1:
        suiid=suiid-1
        uiid=str(suiid)
        sui()
        userui(uiid)


def baocun1():  ###s
    global usernicheng
    print('1')
    if scodeentry.get()!='':
        newcode=scodeentry.get()
        newcode =encrypt1(newcode,logcode)
        udps.sendto(('0018'+userid+newcode).encode('gb2312'), ip)
    if snichengentry.get() != ws['B1'].value:
        usernicheng = snichengentry.get()
        udps.sendto(('0017' + userid + usernicheng).encode('gb2312'), ip)
        ws['B1'].value= usernicheng
        nicheng.config(text=usernicheng)
    udps.sendto(('0019' + userid + uiid).encode('gb2312'), ip)
    window2.destroy()


def suserdelete(event):
        print('1')
        window2.destroy()
        try:
            os.remove(userid + ".xlsx")
        except:
            pass
        userexit1(event)










def getnew():  #请求新消息

    while True:
     time.sleep(3)
     if userid!='':
      udp('0005'+userid)





getnewt3 = threading.Thread(target=getnew)
getnewt3.setDaemon(True)





def dengluok():          #登录成功事件


    global ws
    global wb

    try:
        wb = load_workbook(userid+".xlsx")
    except:
        wb = Workbook()  # 新建表格()
        ws = wb.active  # 赋值当前活动sheet
        #print(ws.title)  # 打印ws的sheet的标题
        ws.title = "Sheet1"
        ws["A1"].value=userid

    else:
        wb.active = wb["Sheet1"]
        ws = wb.active
    try:
     getnewt3.start()
    except:
        print('1')
    else:
        print('1')
    userinput.place_forget()
    codeinput.place_forget()
    denglu.place_forget()
    codeshow.place_forget()
    codeshow.pack_forget()
    global nicheng
    global uiid
    nicheng=tk.Label(window)
    nicheng.config(text=' ')
    udp('0009' + userid)




    friendline1()
    global newchatlable
    global newchatentry
    global userexit
    global shezhilabel
    global useridlabel
    window.geometry('800x530+400+200')
    windowmin.place(x=725,y=2)
    windowoff.place(x=755,y=2)
    nicheng.place(x=10, y=40)
    udpsituation.place(x=110, y=40)
    newchatlable=tk.Label(window,text='发起新聊天')
    newchatentry=tk.Entry(window,text='请输入对方账号',bd=0)
    useridlabel=tk.Label(window,text=userid)
    shezhilabel = tk.Label(window, text='设置')
    userexit=tk.Label(window,text='注销')
    newchatentry.insert(END,'请输入对方账号')
    newchatentry.bind('<FocusIn>',newchatentryfocusin)
    newchatentry.bind('<FocusOut>',newchatentryfocusout)
    newchatentry.bind("<Return>",newchat)
    userexit.bind("<Button-1>",userexit1)
    shezhilabel.bind("<Button-1>",shezhi)
    newchatlable.place(x=10,y=100)
    newchatentry.place(x=90, y=102)
    userexit.place(x=70, y=70)
    useridlabel.place(x=10, y=70)
    shezhilabel.place(x=110,y=70)





def userui(uiid):
    print('uiid为'+uiid)
    if int(uiid)<=2:
        uimain='1'
    else:uimain='2'
    global indexbgimage
    global offimage0
    global minimage0
    global offimage1
    global minimage1
    global useridlabel
    indexbgimage = tk.PhotoImage(file="组件/ui/"+uimain+"/用户背景"+uiid+".png")
    offimage0 = tk.PhotoImage(file="组件/ui/" + uimain + "/关闭0.png")
    minimage0 = tk.PhotoImage(file="组件/ui/" + uimain + "/最小化0.png")
    offimage1 = tk.PhotoImage(file="组件/ui/" + uimain + "/关闭1.png")
    minimage1 = tk.PhotoImage(file="组件/ui/" + uimain + "/最小化1.png")
    indexbgi.config(image=indexbgimage)
    if uimain=='1':
        indexbgi.config(bg='black')
        windowoff.config(image=offimage0,bg='black')
        windowmin.config(image=minimage0,bg='black')
        userinput.config(bg='#555555')
        codeinput.config(bg='#555555')
        udpsituation.config(fg='#8E8E84',bg='black')
        result.config(bg='#2B2B2B',fg='#FFFFFF')
        setput.config(bg='#535353',fg='#FFFFFF')
        codeshow.config(bg='black',fg='#8E8E84')
        nicheng.config(bg='black',fg='#8E8E84')
        newchatlable.config(bg='black', fg='#8E8E84')
        useridlabel.config(bg='black', fg='#8E8E84')
        shezhilabel.config(bg='black', fg='#8E8E84')
        userexit.config(bg='black',fg='#8E8E84')
        newchatentry.config(bg='#535353',fg='#8E8E84')
        friendlist1.config(bg='#2B2B2B',fg='#8E8E84')
        friendlist2.config(bg='#2B2B2B', fg='#8E8E84')
        friendlist3.config(bg='#2B2B2B', fg='#8E8E84')
        friendlist4.config(bg='#2B2B2B', fg='#8E8E84')
        friendlist5.config(bg='#2B2B2B', fg='#8E8E84')
        friendlist6.config(bg='#2B2B2B', fg='#8E8E84')
        friendlist7.config(bg='#2B2B2B', fg='#8E8E84')
        friendlist8.config(bg='#2B2B2B', fg='#8E8E84')
        window.config(bg='black')
    if uimain=='2':
        indexbgi.config(bg='white')
        windowoff.config(image=offimage0,bg='white')
        windowmin.config(image=minimage0,bg='white')
        userinput.config(bg='#F0F0F0')
        codeinput.config(bg='#F0F0F0')
        udpsituation.config(fg='#8E8E84',bg='white')
        result.config(bg='#C1C1C1',fg='#747B74')
        setput.config(bg='#EBECDE',fg='#747B74')
        codeshow.config(bg='white',fg='#8E8E84')
        nicheng.config(bg='white',fg='#8E8E84')
        newchatlable.config(bg='white', fg='#8E8E84')
        useridlabel.config(bg='white', fg='#8E8E84')
        shezhilabel.config(bg='white', fg='#8E8E84')
        userexit.config(bg='white',fg='#8E8E84')
        newchatentry.config(bg='#EBECDE',fg='#8E8E84')
        friendlist1.config(bg='#EBECDE',fg='#8E8E84')
        friendlist2.config(bg='#EBECDE', fg='#8E8E84')
        friendlist3.config(bg='#EBECDE', fg='#8E8E84')
        friendlist4.config(bg='#EBECDE', fg='#8E8E84')
        friendlist5.config(bg='#EBECDE', fg='#8E8E84')
        friendlist6.config(bg='#EBECDE', fg='#8E8E84')
        friendlist7.config(bg='#EBECDE', fg='#8E8E84')
        friendlist8.config(bg='#EBECDE', fg='#8E8E84')
        window.config(bg='white')












def userexit1(event):
    global indexbgimage
    global offimage0
    global minimage0
    global offimage1
    global minimage1
    global useridlabel
    userui('1')
    codeshow.deselect()
    window.geometry('510x258+400+200')
    indexbgimage = tk.PhotoImage(file="组件/ui/1/登录背景1.png")
    indexbgi.config(image=indexbgimage)


    userinput.place(x=300, y=50)
    codeinput.place(x=300, y=85)
    denglu.place(x=300, y=150)
    codeshow.place(x=400, y=113)

    udpsituation.place(x=310, y=115)
    windowoff.place(x=465, y=2)
    windowmin.place(x=435, y=2)
    nicheng.place_forget()

    useridlabel.place_forget()
    shezhilabel.place_forget()
    newchatlable.place_forget()
    newchatentry.place_forget()
    friendlist1.place_forget()
    friendlist2.place_forget()
    friendlist3.place_forget()
    friendlist4.place_forget()
    friendlist5.place_forget()
    friendlist6.place_forget()
    friendlist7.place_forget()
    friendlist8.place_forget()
    result.place_forget()
    setput.place_forget()
    userexit.place_forget()
    newchatentry.delete(0, END)
    global userid
    global nowfriendid
    nowfriendid=''
    userid=''



def newchatentryfocusin(event):
    global newchatlable
    global newchatentry
    if newchatentry.get()=='请输入对方账号':
        newchatentry.delete(0, END)
def newchatentryfocusout(event):
    global newchatlable
    global newchatentry
    if newchatentry.get() == '':
        newchatentry.insert(END,'请输入对方账号')



def newchat(event):
    global newchatlable
    global newchatentry
    global newchatlablesendend
    udps.sendto(('0011'+str(newchatentry.get())+'&'+userid+'&'+time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())+'&'+'你好').encode('gb2312'), ip)
    #print('0011'+str(newchatentry.get())+userid+'&'+time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())+'&'+'你好')
    newchatlablesendend = tk.Label(window, text='请求已发送,等待对方回信')
    newchatlablesendend.place(x=400, y=850)
    newchatentry.delete(0, END)











    

def friendline1():
    global friendlist1
    global friendlist2
    global friendlist3
    global friendlist4
    global friendlist5
    global friendlist6
    global friendlist7
    global friendlist8
    global firstlist
    global friendlistmainy
    global friendlistmainymin
    global friendlistmainsum
    global friendlistmainx
    global friendlistmain
    global friendline
    global friendlinetime
    friendline=[]
    friendlinetime = []
    friendlinetime1 = []
    for col in ws.iter_cols(min_row=2, min_col=1, max_row=999, max_col=1):
        for cell in col:
            if cell.value!=None:
                friendlinetime1.append(ws["C"+str(cell.row)].value)
            else:break
    friendlinetime=sorted(friendlinetime1,reverse=True)
    i=0
    while i<len(friendlinetime):
        for col in ws.iter_cols(min_row=2, min_col=3, max_row=999, max_col=3):
            for cell in col:
                if cell.value == friendlinetime[i]:

                    friendline.append(ws["A" + str(cell.row)].value)
                if cell.value==None:
                    break
        i=i+1
    i=0
    print(friendline)
    print(friendlinetime)
    friendlist1 = tk.Label(window, height=2, bg='blue', width=40, text='1')
    friendlist2 = tk.Label(window, height=2, bg='blue', width=40, text='2')
    friendlist3 = tk.Label(window, height=2, bg='blue', width=40, text='3')
    friendlist4 = tk.Label(window, height=2, bg='blue', width=40, text='4')
    friendlist5 = tk.Label(window, height=2, bg='blue', width=40, text='5')
    friendlist6 = tk.Label(window, height=2, bg='blue', width=40, text='6')
    friendlist7 = tk.Label(window, height=2, bg='blue', width=40, text='7')
    friendlist8 = tk.Label(window, height=2, bg='blue', width=40, text='8')


    friendlist1.bind(" <Button-1>",friendlist1click)
    friendlist2.bind(" <Button-1>", friendlist2click)
    friendlist3.bind(" <Button-1>", friendlist3click)
    friendlist4.bind(" <Button-1>", friendlist4click)
    friendlist5.bind(" <Button-1>", friendlist5click)
    friendlist6.bind(" <Button-1>", friendlist6click)
    friendlist7.bind(" <Button-1>", friendlist7click)
    friendlist8.bind(" <Button-1>", friendlist8click)

    firstlist = friendlist1

    friendlistmainymin = 170
    friendlistmain = 1
    friendlistmainy = friendlistmainymin
    friendlistmainx=15

    if len(friendline)>=9:
        global friendlist8id
        global friendlist1id
        global friendlist2id
        global friendlist3id
        global friendlist4id
        global friendlist5id
        global friendlist6id
        global friendlist7id
        friendlist1id = 1
        friendlist2id = 2
        friendlist3id = 3
        friendlist4id = 4
        friendlist5id = 5
        friendlist6id = 6
        friendlist7id = 7
        friendlist8id = 8
        friendlistmainsum = 1
        friendlist1.config(text=str(friendline[friendlist1id - 1]))
        friendlist2.config(text=str(friendline[friendlist2id - 1]))
        friendlist3.config(text=str(friendline[friendlist3id - 1]))
        friendlist4.config(text=str(friendline[friendlist4id - 1]))
        friendlist5.config(text=str(friendline[friendlist5id - 1]))
        friendlist6.config(text=str(friendline[friendlist6id - 1]))
        friendlist7.config(text=str(friendline[friendlist7id - 1]))
        friendlist8.config(text=str(friendline[friendlist8id - 1]))
        friendlistmove()
        window.bind('<MouseWheel>', mousemove)
    else:
        friendlistcalm()

def friendlistcalm():
    global friendlist1
    global friendlist2
    global friendlist3
    global friendlist4
    global friendlist5
    global friendlist6
    global friendlist7
    global friendlist8
    global firstlist
    global friendlistmainy
    global friendlistmain
    global friendline
    global friendlistmainx
    global friendlinetime
    global friendlistmain
    global friendlistmainsum
    global friendlist8id
    global friendlist1id
    global friendlist2id
    global friendlist3id
    global friendlist4id
    global friendlist5id
    global friendlist6id
    global friendlist7id
    friendlist1id=1
    friendlist2id = 2
    friendlist3id = 3
    friendlist4id = 4
    friendlist5id = 5
    friendlist6id = 6
    friendlist7id = 7
    friendlist8id = 8
    if len(friendline)>=1:
        #friendlist1 = tk.Label(window, height=2, bg='blue', width=40, text='1')
        friendlist1.config(text=str(friendline[friendlist1id-1]))
        friendlist1.place(x=friendlistmainx, y=friendlistmainy)
        if len(friendline) >= 2:
            #friendlist2 = tk.Label(window, height=2, bg='blue', width=40, text='2')
            friendlist2.config(text=str(friendline[friendlist2id-1]))
            friendlist2.place(x=friendlistmainx, y=friendlistmainy+45*1)
            if len(friendline) >= 3:
                #friendlist3 = tk.Label(window, height=2, bg='blue', width=40, text='3')
                friendlist3.config(text=str(friendline[friendlist3id-1]))
                friendlist3.place(x=friendlistmainx, y=friendlistmainy+45*2)
                if len(friendline) >= 4:
                    #friendlist4 = tk.Label(window, height=2, bg='blue', width=40, text='4')
                    friendlist4.config(text=str(friendline[friendlist4id-1]))
                    friendlist4.place(x=friendlistmainx, y=friendlistmainy+45*3)
                    if len(friendline) >= 5:
                        #friendlist5 = tk.Label(window, height=2, bg='blue', width=40, text='5')
                        friendlist5.config(text=str(friendline[friendlist5id-1]))
                        friendlist5.place(x=friendlistmainx, y=friendlistmainy+45*4)
                        if len(friendline) >= 6:
                            #friendlist6 = tk.Label(window, height=2, bg='blue', width=40, text='6')
                            friendlist6.config(text=str(friendline[friendlist6id-1]))
                            friendlist6.place(x=friendlistmainx, y=friendlistmainy+45*5)
                            if len(friendline) >= 7:
                                #friendlist7 = tk.Label(window, height=2, bg='blue', width=40, text='7')
                                friendlist7.config(text=str(friendline[friendlist7id-1]))
                                friendlist7.place(x=friendlistmainx, y=friendlistmainy+45*6)
                                if len(friendline) >= 8:
                                    #friendlist8 = tk.Label(window, height=2, bg='blue', width=40, text='8')
                                    friendlist8.config(text=str(friendline[friendlist8id-1]))
                                    friendlist8.place(x=friendlistmainx, y=friendlistmainy+45*1)

       
    
    













def addnicheng(nicheng1):    #为新用户添加昵称 #修改本地昵称
    global usernicheng
    print(nicheng)
    ws["B1"].value=nicheng1
    usernicheng=nicheng1
    nicheng.config(text=ws["B1"].value)
    wb.save(userid + ".xlsx")



def nouser(): ##登录无此用户事件
    a=1
def codeerror(): ##登录密码错误事件
    a = 1

def udp(senddate):  #数据发送与监听事件



    if udpsituation1==1 or udpsituation1 == 0:
       print(senddate)
       udps.sendto(senddate.encode('gb2312'), ip)  # 发送数据客户端ip




receive=1
def udpreceive():
 global receive
 while receive==1:


    global code
    global userid
    global m
    try:
     receivedate = udps.recvfrom(10240)  # 等待接受(数据最大为1024)
    except:
        continue
    else:
     指令 = receivedate[0][0:4].decode('gb2312')
     print(指令)
     if 指令=='0002':
           code=receivedate[0][4:].decode('gb2312')
           code=bytes(code, encoding='utf-8')
           userid=userinput.get()
           dengluok()
     if 指令=='0003':
           nouser()
     if 指令=='0004':                                                                   #未进行
           codeerror()
     if 指令=='0010':
           print(receivedate[0][4:])
           addnicheng(receivedate[0][4:])
           udp('0015'+userid)
     if 指令=='0014':
           global friendnicheng
           friendnicheng=receivedate[0][4:]
           friendlistclick2()
     if 指令 == '0016':
           global uiid

           uiid1=receivedate[0][4:]
           uiid1=str(uiid1)
           if uiid1[0]=='b':
               uiid1=uiid1[2:]
               uiid1 = uiid1[:-1]
           ws['C1'].value = uiid1
           if uiid1!=uiid:
               wb.save(userid + ".xlsx")
           uiid=uiid1
           uiid = ws["C1"].value
           print('uiid'+uiid)
           userui(uiid1)
     if 指令 == '0006':
           m = 0.001
           receivedate1=receivedate[0].decode('gb2312')
           #print(receivedate1)
           i = 4
           list = ['123']
           datebegin = 4
           global dateend
           dateend=0
           while i <= len(receivedate1)-1:
               if receivedate1[i] == '&':
                   dateend = i

                   list.append(receivedate1[datebegin:dateend])

                   datebegin = i + 1
               i = i + 1
           list.append(receivedate1[datebegin:])
           i = 1
           mm = 0
           sum1 = 0

           if list[1]!='':
             while i <= len(list)-1:

               date = decrypt1(list[i], code)
               friendid = date[0:5]

               while mm<len(date):
                   if date[mm]=='&' :
                       if sum1==0:
                           sum1=1
                           datetimebegin=mm

                       elif sum1==1:
                           datetime=date[datetimebegin+1:mm]
                           date=date[mm+1:]
                           sum1=0
                           break
                   mm=mm+1
               mm = 0
               sum1 = 0


               date=datetime+'&'+friendid+date

               date = encrypt1(date, code)

               for col in ws.iter_cols(min_row=2, min_col=1, max_row=999, max_col=1):
                   for cell in col:
                       cellrow = cell.row
                       #print(friendid)
                       #print(cell.value)
                       if cell.value == friendid:

                           print('检索到用户')
                           for row in ws.iter_rows(min_row=cellrow, min_col=11, max_row=cellrow,
                                                   max_col=999):  # 括号代表遍历第一行到第二行,第二列到第三列
                               for cell in row:
                                   if cell.value == None:
                                       cell.value = date
                                       ws["C"+str(cell.row)].value=time.time()
                                       ws["C" + str(cell.row)].value =ws["C"+str(cell.row)].value-m
                                       if nowfriendid != '':
                                           friendlistclick()

                                       print(time.time() - m)
                                       print(m)

                                       i = i + 1
                                       m = m+0.001


                                       break
                               break
                           break
                       elif cell.value == None:
                           print('未检索到用户')
                           cell.value = friendid
                           ws["K" + str(cell.row)].value = date
                           ws["C" + str(cell.row)].value = time.time()
                           ws["C" + str(cell.row)].value = ws["C" + str(cell.row)].value-m
                           if nowfriendid != '':
                               friendlistclick()

                           #print(m)
                           m = m + 0.001
                           break
             wb.save(userid + ".xlsx")
             m=0.001

           print("nh" + str(len(list)))


     global udpsituation1

 return
udpreceivet4=threading.Thread(target=udpreceive)
udpreceivet4.setDaemon(True)
def denglu1():   #
    user=userinput.get()
    code=codeinput.get()
    code=encrypt1(code,logcode)
    if user!='' and user!='Pansis账户' and user.isnumeric():
       a="0001"+userinput.get()+'&'+code
       udp(a)
    return
def userinputdelte(event):
    if userinput.get()=='Pansis账户':
        userinput.delete(0, END)
    return
def userinputadd(event):
    if userinput.get()=='':
        userinput.insert(END,'Pansis账户')
    return
def codeinputdelte(event):
    if codeinput.get()=='密码':
        codeinput.delete(0, END)
        codeshow1()
    return
def codeinputadd(event):
    if codeinput.get()=='':
        codeinput.insert(END,'密码')
        codeinput.config(show='')
    return


def windowoff1(event):  ###
    print('1')
    sys.exit(1)
    window.destroy()


def windowmin1(event):    ###
    print('2')
    #window.state('icon')
def windowoff2(event):
    windowoff.config(image=offimage1)
def windowmin2(event):
    windowmin.config(image=minimage1)
def windowoff3(event):
    windowoff.config(image=offimage0)
def windowmin3(event):
    windowmin.config(image=minimage0)

def windowmove(event):
    new_x = (event.x - window.x) + window.winfo_x()
    new_y = (event.y - window.y) + window.winfo_y()
    s = f"{window.winfo_width()}x{window.winfo_height()}+{new_x}+{new_y}"
    window.geometry(s)
def windowclick(event):
        """获取当前窗口位置并保存"""
        window.x, window.y = event.x, event.y


denglu=tk.Button(window,bg='#298CF4',relief=FLAT,width=25,command=denglu1,text='登录',fg='white',bd=0,activebackground='#298CF4')
codeshow=tk.Checkbutton(window,text='显示密码',command=codeshow1,variable=codeshowvariable,bg='black',fg='#8E8E84')
denglu.place(x=300,y=150)
codeshow.deselect()
codeshow.place(x=400,y=113)
userinput.bind("<FocusOut>",userinputadd)
userinput.bind("<FocusIn>",userinputdelte)
userinput.insert(END,'Pansis账户')
codeinput.bind("<FocusOut>",codeinputadd)
codeinput.bind("<FocusIn>",codeinputdelte)
#setput.bind("<Return>",liaotiansend)
codeinput.insert(END,'密码')
windowoff.bind("<Button-1>",windowoff1)
windowmin.bind("<Button-1>",windowmin1)
udpsitiontest()
windowoff.bind("<Enter>",windowoff2)
windowoff.bind("<Leave>",windowoff3)
windowmin.bind("<Enter>",windowmin2)
windowmin.bind("<Leave>",windowmin3)
indexbgi.bind("<B1-Motion>",windowmove)
indexbgi.bind("<Button-1>",windowclick)
udpreceivet4.start()
window.mainloop()











udps.close                       #关闭套接字