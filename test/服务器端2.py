# coding=utf-8
# This is a sample Python script.
# !/usr/bin/python
# -*- coding: UTF-8 -*-
#!/usr/bin/env python

import tkinter as tk
from tkinter import FLAT
from socket import socket,AF_INET,SOCK_DGRAM
import time
import os
from cryptography.fernet import Fernet
from openpyxl import Workbook
from openpyxl import load_workbook
import random                 #引用区

udps=socket(AF_INET,SOCK_DGRAM)   #创建套接字
udps.bind(('',8080))   #本机软件绑定端口
logcode=b'w2V8gAUFER-9uq6lKG7AfpYI6AcINEs7Ipm4KyQj2AI='  #登录密码秘钥


wb=load_workbook("server.xlsx")
wb.active=wb["Sheet1"]
ws=wb.active
on=1




def encrypt1(date,cipher_key):     # 进行加密

    date = bytes(date, encoding='utf-8')

    encrypted_text = Fernet(cipher_key).encrypt(date)

    encrypted_text = str(encrypted_text, encoding='utf-8')
    return encrypted_text


def decrypt1(date, cipher_key):  # 进行解密
    date = bytes(date, encoding='utf-8')
    decrypted_text = Fernet(cipher_key).decrypt(date)

    decrypted_text = str(decrypted_text, encoding='utf-8')
    return decrypted_text




def udp(senddate,ip):              #用于数据发送
    udps.sendto(senddate.encode('gb2312'),(ip,8081) )  # 发送数据客户端ip
    #receivedate = udps.recvfrom(1024)  # 等待接受(数据最大为1024)
    return
def ser0001(date,ip):            #返回登录结果
    #print(date)
    i=0
    while i<=len(date)-1:
        a=date[i]

        if a=='&':
            user=date[0:i]
            code=date[i+1:]
            code=decrypt1(code,logcode)
            print(user)
            print(code)
        i=i+1
    i=0
    for col in ws.iter_cols(min_row=1, min_col=1, max_row=99, max_col=1):
        for cell in col:

            if cell.value==user:

                if str(ws['B'+str(cell.row)].value)==code:
                    dete="0002"+ws['C'+str(cell.row)].value
                    udp(dete,ip)
                    break
                else:
                    udp("0004", ip)
                    break
            if cell.value==None:
                udp("0003", ip)
                break
def ser0005(date,ip):
    global datesum
    i=0
    while i<=len(date)-1:
        if date[i]!='0':
            datesum=int(date[i:])
            break
        i=i+1
    i=11
    date='0006'
    while i>1:
        if ws.cell(row=datesum, column=i).value==None:
            i=0
            print('OK')
            print(date)
            udps.sendto(date.encode('gb2312'), (ip, 8081))



        else:
            if date!='0006':
                date = date + '&' + str(ws.cell(row=datesum, column=i).value)
                print(date)
                ws.cell(row=datesum, column=i).value = None
            else:
                date=date+str(ws.cell(row=datesum, column=i).value)
                ws.cell(row=datesum, column=i).value = None

            i = i + 1

    wb.save("server.xlsx")






def ser0009(date,ip):
    for col in ws.iter_cols(min_row=1, min_col=1, max_row=99, max_col=1):
        for cell in col:
            print(ws['D' + str(cell.row)].value)
            if cell.value==date:

                udp('0010'+ws['D'+str(cell.row)].value, ip)
                break
def ser0013(date,ip):
    for col in ws.iter_cols(min_row=1, min_col=1, max_row=99, max_col=1):
        for cell in col:
            print(ws['D' + str(cell.row)].value)
            if cell.value==date:

                udp('0014'+ws['D'+str(cell.row)].value, ip)
                break
def ser0011(date,ip):
    receiveid=date[0:5]
    sendid=date[6:11]
    print('接收人'+receiveid)
    print('发送人'+sendid)

    date=sendid+date[11:]
    for col in ws.iter_cols(min_row=1, min_col=1, max_row=999, max_col=1):
        for cell in col:
            if cell.value==receiveid:
                cellrow=cell.row
                code=ws["C"+str(cell.row)].value
                code = bytes(code, encoding='utf-8')
                date=encrypt1(date,code)
                for row in ws.iter_rows(min_row=cellrow, min_col=11, max_row=cellrow,
                                        max_col=999):  # 括号代表遍历第一行到第二行,第二列到第三列
                    for cell in row:
                        if cell.value == None:
                            cell.value = date
                            wb.save('server.xlsx')
                            break

                break
            if cell.value==None:
                break

def ser0015(date,ip):
    for col in ws.iter_cols(min_row=1, min_col=1, max_row=999, max_col=1):
        for cell in col:
            if cell.value==date:
                udp('0016' + str(ws["E"+str(cell.row)].value), ip)
                print('123')
                break
            if cell.value==None:
                udp('0000', ip)
                print('1234')
                break

def ser0018(date,ip):
    userid=date[0:5]
    date=date[5:]
    print('用户'+userid)
    date=decrypt1(date,logcode)
    for col in ws.iter_cols(min_row=1, min_col=1, max_row=999, max_col=1):
        for cell in col:
            if cell.value==userid:

                ws["B"+str(cell.row)]=date
                wb.save('server.xlsx')
                break
            if cell.value==None:
                break



def ser0017(date,ip):
    userid=date[0:5]
    date=date[5:]
    print('用户'+userid)
    for col in ws.iter_cols(min_row=1, min_col=1, max_row=999, max_col=1):
        for cell in col:
            if cell.value==userid:
                ws["D"+str(cell.row)]=date
                wb.save('server.xlsx')
                break
            if cell.value==None:
                break

def ser0019(date,ip):
    userid=date[0:5]
    date=date[5:]
    print('用户'+userid)
    for col in ws.iter_cols(min_row=1, min_col=1, max_row=999, max_col=1):
        for cell in col:
            if cell.value==userid:
                ws["E"+str(cell.row)]=date
                wb.save('server.xlsx')
                break
            if cell.value==None:
                break












while on==1:
    try:
         receivedate = udps.recvfrom(1024)  # 等待接受(数据最大为1024)
    except:
        continue
    else:
          nowtime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())  # 格式化成2016-03-20 11:45:39形式
          print(receivedate[0].decode('gb2312'))
    #print('来自%s<%s> \n %s' % (receivedate[1], nowtime, receivedate[0].decode('gb2312')))
          指令=receivedate[0][0:4].decode('gb2312')
          ip=receivedate[1][0]
          print(receivedate[1][0])
          if 指令=='0001':
                ser0001(receivedate[0][4:].decode('gb2312'),ip)
          if 指令=='0000':
                udp("0005", ip)
          if 指令=='0005':
                ser0005(receivedate[0][4:].decode('gb2312'), ip)
          if 指令=='0009':
                ser0009(receivedate[0][4:].decode('gb2312'), ip)
          if 指令=='0013':
                ser0013(receivedate[0][4:].decode('gb2312'), ip)
          if 指令=='0011':
              ser0011(receivedate[0][4:].decode('gb2312'), ip)
          if 指令 == '0015':
              ser0015(receivedate[0][4:].decode('gb2312'), ip)
          if 指令 == '0018':
              ser0018(receivedate[0][4:].decode('gb2312'), ip)
          if 指令 == '0017':
              ser0017(receivedate[0][4:].decode('gb2312'), ip)
          if 指令 == '0019':
              ser0019(receivedate[0][4:].decode('gb2312'), ip)




    #print(指令)




udps.close
