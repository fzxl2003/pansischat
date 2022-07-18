# coding=utf-8
# !/usr/bin/python
# -*- coding: UTF-8 -*-
# !/usr/bin/env python

# import tkinter as tk
# from tkinter import FLAT
import hashlib
from socket import *
import time
# import os
from cryptography.fernet import Fernet
# from openpyxl import Workbook
from openpyxl import load_workbook
from threading import Thread
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.asymmetric import padding
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives import serialization
from cryptography.hazmat.primitives.asymmetric import rsa
import fileserver

# logcode=b'w2V8gAUFER-9uq6lKG7AfpYI6AcINEs7Ipm4KyQj2AI='  #登录密码秘钥


wb = load_workbook("server.xlsx")
wb.active = wb["Sheet1"]
ws = wb.active
on = 1

userwithip = {}  # ip与登录的用户的字典查询


def encrypt1(data, cipher_key):  # 进行加密

    data = bytes(data, encoding='utf-8')

    encrypted_text = Fernet(cipher_key).encrypt(data)

    encrypted_text = str(encrypted_text, encoding='utf-8')
    return encrypted_text


def decrypt1(data, cipher_key):  # 进行解密
    if data[:2] == 'b\'':
        print('123')
        data = data[2:]
        data = data[:-1]
        print(data)
    data = bytes(data, encoding='utf-8')
    print(data)
    print(cipher_key)
    decrypted_text = '1234'
    try:
        decrypted_text = Fernet(cipher_key).decrypt(data)
        decrypted_text = str(decrypted_text, encoding='utf-8')
    except:

        print(decrypted_text)
    return decrypted_text


def ser0001(data, clientinfo):  # 返回登录结果
    global userwithip
    # print(data)
    i = 0
    user=''
    code=''
    returndata = '0003'
    while i <= len(data) - 1:
        a = data[i]

        if a == '&':
            user = data[0:i]
            code = data[i + 1:]
            # code=decrypt1(code,logcode)
            print(user)
            print(code)
            break
        i = i + 1
    i = 0
    for col in ws.iter_cols(min_row=1, min_col=1, max_row=99, max_col=1):
        for cell in col:
            print(cell.value)

            if cell.value == user:
                rightcode = str(ws['B' + str(cell.row)].value)
                if rightcode[-1:] != "&":
                    # rightcode = decrypt1(rightcode,ws['C'+str(cell.row)].value)
                    rightcode = hashlib.md5(rightcode.encode('utf8')).hexdigest()
                    ws['B' + str(cell.row)].value = rightcode + '&'
                    wb.save("server.xlsx")
                    print(rightcode)
                else:
                    rightcode = rightcode[:-1]
                if rightcode == code:
                    usercode = str(ws['C' + str(cell.row)].value)
                    data = "0002" + str(usercode) + "*" + str(ws['D' + str(cell.row)].value) + "*" + str(
                        ws['E' + str(cell.row)].value)
                    userwithip[str(clientinfo[0]) + str(clientinfo[1])] = user
                    data2 = '0006'
                    data = data + "*" + data2
                    returndata = data
                    break
                else:
                    returndata = "0004"

                    break
            if cell.value == None:
                returndata = "0003"
                break
    print(returndata)
    return returndata


def ser0005(data):  # 仅用于整理新消息
    global datasum
    i = 0
    returndata=''
    while i <= len(data) - 1:
        if data[i] != '0':
            datasum = int(data[i:])
            break
        i = i + 1
    i = 11
    data = '0006'
    while i > 1:
        if ws.cell(row=datasum, column=i).value == None:
            i = 0
            # print('OK')
            returndata = data
        else:

            if data != '0006':

                data = data + '&' + str(ws.cell(row=datasum, column=i).value)

                print("abcd" + data)
                ws.cell(row=datasum, column=i).value = None
            else:

                data = data + str(ws.cell(row=datasum, column=i).value)
                print("efg" + data)
                ws.cell(row=datasum, column=i).value = None
            i = i + 1
    wb.save("server.xlsx")
    # print("获取的新消息为："+returndata)
    return returndata


# def ser0009(data):
#  for col in ws.iter_cols(min_row=1, min_col=1, max_row=99, max_col=1):
#      for cell in col:
#          print(ws['D' + str(cell.row)].value)
#         if cell.value==data:
#             returndata ='0010'+ws['D'+str(cell.row)].value
#              break
#  return returndata
def ser0013(data):
    returndata = "OO14未知"
    for col in ws.iter_cols(min_row=1, min_col=1, max_row=99, max_col=1):
        for cell in col:
            print(ws['D' + str(cell.row)].value)
            if cell.value == data:
                returndata = '0014' + ws['D' + str(cell.row)].value
                break
    return returndata


def ser0011(data):
    receiveid = data[0:5]
    sendid = data[6:11]
    print('接收人' + receiveid)
    print('发送人' + sendid)
    data = data[11:]
    for col in ws.iter_cols(min_row=1, min_col=1, max_row=999, max_col=1):
        for cell in col:
            if cell.value == sendid:
                data = decrypt1(data, ws['C' + str(cell.row)].value)

    data = sendid + '&' + data
    for col in ws.iter_cols(min_row=1, min_col=1, max_row=999, max_col=1):
        for cell in col:
            if cell.value == receiveid:
                cellrow = cell.row
                code = ws["C" + str(cell.row)].value
                code = bytes(code, encoding='utf-8')
                data = encrypt1(data, code)
                for row in ws.iter_rows(min_row=cellrow, min_col=11, max_row=cellrow,
                                        max_col=999):  # 括号代表遍历第一行到第二行,第二列到第三列
                    for cell in row:
                        if cell.value == None:
                            cell.value = data
                            wb.save('server.xlsx')
                            break

                break
            if cell.value == None:
                break


# def ser0015(data):
#   for col in ws.iter_cols(min_row=1, min_col=1, max_row=999, max_col=1):
#      for cell in col:
#          if cell.value==data:
#              returndata ='0016' + str(ws["E"+str(cell.row)].value)
#               print('123')
#               break
#           if cell.value==None:
##            returndata ='0000'
#              print('1234')
#             break
# return returndata


def ser0017(data):
    userid = data[4:9]
    i = 9
    typenum = 1
    end=0
    usernicheng=''
    denglucode=''
    uiid=''
    while i <= len(data) - 1:
        if data[i] == "*":
            if typenum == 1:
                usernicheng = data[9:i]
            if typenum == 2:
                uiid = data[end + 1:i]
                denglucode = data[i + 1:]
                break
            end = i
            typenum = typenum + 1
        i = i + 1
    print('用户' + userid)
    for col in ws.iter_cols(min_row=1, min_col=1, max_row=999, max_col=1):
        for cell in col:
            if cell.value == userid:
                if usernicheng != "&":
                    ws["D" + str(cell.row)] = usernicheng
                ws["E" + str(cell.row)] = uiid
                if denglucode != "&":
                    ws["B" + str(cell.row)] = denglucode
                wb.save('server.xlsx')
                break
            if cell.value == None:
                break


def ser0019(data):
    userid = data[0:5]
    data = data[5:]
    print('用户' + userid)
    for col in ws.iter_cols(min_row=1, min_col=1, max_row=999, max_col=1):
        for cell in col:
            if cell.value == userid:
                ws["E" + str(cell.row)] = data
                wb.save('server.xlsx')
                break
            if cell.value == None:
                break





# def ser0020():
#   a=1
# t1 = Thread(target=fileserver.start,args=(12345,a))
# t1.start()
#  return
def decryptplus1(receivedata, server_private_key2):
    try:
        original_message = server_private_key2.decrypt(receivedata,
                                                       padding.OAEP(mgf=padding.MGF1(algorithm=hashes.SHA256()),
                                                                    algorithm=hashes.SHA256(), label=None))
    except:
        return 'fail'
    else:
        print(original_message)
        return str(original_message, encoding='utf-8')


def encryptplus(data, client_public_key1):
    data = bytes(data, encoding='utf-8')

    try:
        encrypted = b''
        if len(data) <= 100:
            encrypted = client_public_key1.encrypt(
                data,
                padding.OAEP(
                    mgf=padding.MGF1(algorithm=hashes.SHA256()),
                    algorithm=hashes.SHA256(),
                    label=None
                )
            )
        while len(data) > 100:
            data1 = data[:100]
            encrypting = client_public_key1.encrypt(
                data1,
                padding.OAEP(
                    mgf=padding.MGF1(algorithm=hashes.SHA256()),
                    algorithm=hashes.SHA256(),
                    label=None
                )
            )
            encrypted = encrypted + b'&&&' + encrypting
            data = data[100:]
            if len(data) <= 100:
                encrypting = client_public_key1.encrypt(
                    data,
                    padding.OAEP(
                        mgf=padding.MGF1(algorithm=hashes.SHA256()),
                        algorithm=hashes.SHA256(),
                        label=None
                    )
                )
                encrypted = encrypted + b'&&&' + encrypting
                break
    except:
        # print("???")
        return b'fail'
    else:
        # print("????")
        return encrypted


def decryptplus(receivedata, server_private_key2):
    print(b'abcdef' + receivedata)

    try:
        original_message = b''
        if receivedata[:3] == b'&&&':
            print("a")
            i = 3
            beginnum = 3
            while i < len(receivedata):

                if receivedata[i:i + 3] == b'&&&':
                    print("a")
                    decrypting = receivedata[beginnum:i]
                    original_message1 = server_private_key2.decrypt(
                        decrypting,
                        padding.OAEP(
                            mgf=padding.MGF1(algorithm=hashes.SHA256()),
                            algorithm=hashes.SHA256(),
                            label=None
                        )
                    )
                    print("b")
                    original_message = original_message + original_message1
                    beginnum = i + 3
                    print(original_message)
                i = i + 1
            decrypting = receivedata[beginnum:]
            print(decrypting)
            print("c")
            original_message1 = server_private_key2.decrypt(
                decrypting,
                padding.OAEP(
                    mgf=padding.MGF1(algorithm=hashes.SHA256()),
                    algorithm=hashes.SHA256(),
                    label=None
                )
            )
            print(original_message)
            print("c")
            original_message = original_message + original_message1
        else:
            original_message = server_private_key2.decrypt(
                receivedata,
                padding.OAEP(
                    mgf=padding.MGF1(algorithm=hashes.SHA256()),
                    algorithm=hashes.SHA256(),
                    label=None
                )
            )
    except:
        return 'fail'
    else:
        print(original_message)
        return str(original_message, encoding='utf-8')


def encryptplus1(data, client_public_key1):
    try:
        data = bytes(data, encoding='utf-8')
    except:
        a = 1
    print(data)
    try:
        encrypted = client_public_key1.encrypt(
            data,
            padding.OAEP(
                mgf=padding.MGF1(algorithm=hashes.SHA256()),
                algorithm=hashes.SHA256(),
                label=None
            )
        )
    except:

        return b'fail'
    else:

        return encrypted



threadnum = 0
a = ''


def readmsg(clientsocket, clientinfo):
    global threadnum, nowuploadfilename, nowdownloadclient
    clientsocket.settimeout(100)
    sendsum = 0
    nowuploadfilename[str(clientinfo[0]) + str(clientinfo[1])] = ''
    nowdownloadclient[str(clientinfo[0]) + str(clientinfo[1])] = ''
    server_private_key2=''
    client_public_key1=''

    # print('ok')
    print(clientinfo)
    while True:
        try:
            receivedata = clientsocket.recv(10240)  # 等待接受(数据最大为1024)
            # print(receivedata)
        except:
            clientsocket.close()
            threadnum = threadnum - 1
            print("一线程关闭3")
            break
        else:
            nowtime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())  # 格式化成2016-03-20 11:45:39形式
            if receivedata[0:4] == bytes('0024', encoding='utf-8'):
                client_public_key = receivedata[4:]
                client_public_key1 = serialization.load_pem_public_key(client_public_key, backend=default_backend())
                # print(client_public_key1)
                # print(client_public_key)

                client_public_key2 = serialization.load_pem_public_key(client_public_key, backend=default_backend())
                server_private_key = rsa.generate_private_key(
                    public_exponent=65537,
                    key_size=2048,
                    backend=default_backend()
                )
                server_public_key = server_private_key.public_key()
                # store private key
                server_private_key1 = server_private_key.private_bytes(
                    encoding=serialization.Encoding.PEM,
                    format=serialization.PrivateFormat.PKCS8,
                    encryption_algorithm=serialization.NoEncryption()
                )
                server_public_key1 = server_public_key.public_bytes(
                    encoding=serialization.Encoding.PEM,
                    format=serialization.PublicFormat.SubjectPublicKeyInfo
                )
                server_private_key2 = serialization.load_pem_private_key(
                    server_private_key1,
                    password=None,
                    backend=default_backend()
                )
                clientsocket.send((bytes('0024', encoding='utf-8') + server_public_key1))
                continue

            if not (receivedata[0:4].isdigit()):
                # print('abcdef')
                # print(receivedata)
                receivedata = decryptplus(receivedata, server_private_key2)
                if receivedata == 'fail':
                    clientsocket.close()
                    threadnum = threadnum - 1
                    print("一线程关闭4")
                    break
            else:
                receivedata = receivedata.decode('utf-8')
            # print('67789'+receivedata)
            test = encryptplus('123', client_public_key1)
            if test == b'fail':
                clientsocket.close()
                threadnum = threadnum - 1
                print("一线程关闭5")
                break
            指令 = receivedata[0:4]
            ip = clientinfo
            # print(ip)
            if 指令=='0000':
                clientversion=int(receivedata[4:])
                if (clientversion<serverversion):
                    a='0031'+newversionurl
                    print(a)
                    a = encryptplus(a, client_public_key1)
                    clientsocket.send(a)





            if 指令 == '0001':
                a = ser0001(receivedata[4:], clientinfo)
                print('78392' + a)
                a = encryptplus(a, client_public_key1)
                # if a==b'fail':
                #    clientsocket.close()
                #     threadnum = threadnum - 1
                #      print("一线程关闭")
                #     break
                clientsocket.send(a)

            if 指令 == '0013':
                a = ser0013(receivedata[4:])
                a = encryptplus(a, client_public_key1)
                if a == b'fail':
                    clientsocket.close()
                    threadnum = threadnum - 1
                    print("一线程关闭6")
                    break
                clientsocket.send(a)
            if 指令 == '0011':
                ser0011(receivedata[4:])
            #   if 指令 == '0015':
            #      a=ser0015(receivedata[4:])
            #      print(a)
            #     clientsocket.send(a.encode('utf-8'))

            if 指令 == '0017':
                ser0017(receivedata)
            if 指令 == '0019':
                ser0019(receivedata[4:])
            if 指令 == '0022':
                nowuploadfilename[str(clientinfo[0]) + str(clientinfo[1])] = receivedata[4:]
            if 指令 == '0025':
                ser0001(receivedata[4:], clientinfo)
            if 指令 == '0026':
                print("0022222")
                a = ser0026(receivedata[4:], clientsocket, client_public_key1)
                if (a != 'wait'):
                    print(a)
                    a = encryptplus(a, client_public_key1)
                    if a == b'fail':
                        clientsocket.close()
                        threadnum = threadnum - 1
                        break
                    clientsocket.send(a)

            if str(clientinfo[0]) + str(clientinfo[1]) in userwithip.keys():
                a = ser0005(userwithip[str(clientinfo[0]) + str(clientinfo[1])])
                if a != '0006':
                    print(a)
                    c = encryptplus(a, client_public_key1)
                    print(c)
                    if c == b'fail':
                        clientsocket.close()
                        threadnum = threadnum - 1
                        print("一线程关闭7")
                        break
                    clientsocket.send(c)

            try:
                sit = fileserver.fileuploadsituationend

            except:
                a = ''
            else:
                newfilename = fileserver.newfilename[4:]
                port = fileserver.newfilename[0:4]
                if sit == 1:
                    if nowuploadfilename[str(clientinfo[0]) + str(clientinfo[1])] != '':
                        if nowuploadfilename[str(clientinfo[0]) + str(clientinfo[1])] == newfilename:
                            a = '0021' + fileserver.newfilename
                            print('euiwo' + a)
                            a = encryptplus(a, client_public_key1)
                            if a == b'fail':
                                clientsocket.close()
                                threadnum = threadnum - 1
                                print("一线程关闭1")
                                break
                            clientsocket.send(a)
                            fileserver.fileuploadsituationend = 0
                            nowuploadfilename[str(clientinfo[0]) + str(clientinfo[1])] = ''
                        elif newfilename[0] == '(':
                            i = 1
                            while i < len(newfilename):
                                if newfilename[i] == ')':
                                    newfilename = newfilename[i + 1:]
                                    break
                                i = i + 1
                            if nowuploadfilename[str(clientinfo[0]) + str(clientinfo[1])] == newfilename:
                                a = '0021' + fileserver.newfilename
                                print('euiwo' + a)
                                a = encryptplus(a, client_public_key1)
                                if a == b'fail':
                                    clientsocket.close()
                                    threadnum = threadnum - 1
                                    print("一线程关闭2")
                                    break

                                clientsocket.send(a)
                                nowuploadfilename[str(clientinfo[0]) + str(clientinfo[1])] = ''
                                fileserver.fileuploadsituationend = 0
            a = ''


def downloadconnectwait(waittime, ip, clientsocket):
    i = 0
    while i <= 100:
        try:
            port = fileserver.nowdownloadclientport[ip]
        except:
            a = 1
        else:
            clientsocket.send(('0023' + port).encode('utf-8'))
        i = i + 1
        time.sleep(0.01)
        print(i)


def p2pwaitdelete(userid, a):
    time.sleep(30)
    listlen = len(p2plist)
    i = 0
    while (i < listlen):
        if (p2plist[i][0] == userid):
            del p2plist[i]
        i=i+1


p2plist = []


def ser0026(data, clientsocket, key):
    userdata = []
    print(data)
    userdata.append(data[:5])
    userdata.append(data[5:10])
    begin = 9
    i = 10
    strlen = len(data)
    leibie = 1
    print('fdjskl')
    while (i < strlen):
        if (data[i] == '&'):
            shuju = data[begin+1:i]
            begin = i
            leibie = leibie + 1
            userdata.append(shuju)
            if (leibie == 3):
                userdata.append(data[begin+1:])
        i = i + 1
    for list in p2plist:
        if (list[1] == userdata[0]):
            print('nownow')
            a = '0027' + userdata[2] + '&' + userdata[3] + '&' + userdata[4]
            a = encryptplus(a, list[6])
            if (a != b'fail'):
                list[5].send(a)
            return '0027' + list[2] + '&' + list[3] + '&' + list[4]
    print('nownow111')
    userdata.append(clientsocket)
    userdata.append(key)
    #ser0011(userdata[1] + '&' + userdata[0] + '&' + time.strftime("%Y-%m-%d %H:%M:%S",
                                                                  #time.localtime()) + '&好友请求P2P聊天，请点击P2P按钮接受')
    p2plist.append(userdata)
    print('hjdksa' + userdata[0])
    t = Thread(target=p2pwaitdelete, args=(userdata[0], 1))
    t.start()
    return 'wait'


def udpserver():
    p2ps = socket(AF_INET, SOCK_DGRAM)  # 创建套接字
    p2ps.bind(('0.0.0.0', 7583))

    while (True):
        data, adr = p2ps.recvfrom(1024)
        data = data.decode('utf-8')
        print('21321' + data)
        order = data[0:4]
        data = data[4:]
        if (order == '0015'):
            # print('3242')
            data = '0016' + adr[0] + '&' + str(adr[1])
            print(data)
            p2ps.sendto(bytes(data.encode('utf-8')), adr)


def main():
    global udps
    global userwithip
    global threadnum
    global nowuploadfilename, nowdownloadclient,serverversion,newversionurl
    nowuploadfilename = {}
    nowdownloadclient = {}
    udps = socket(AF_INET, SOCK_STREAM)  # 创建套接字
    udps.bind(('', 8080))  # 本机软件绑定端口
    verisontxt = open('information/version.txt', mode='r')
    serverversion = int (verisontxt.read())
    verisontxt.close()
    print(verisontxt)
    verisontxt = open('information/newversionurl.txt', mode='r')
    newversionurl = verisontxt.read()
    print(newversionurl)
    verisontxt.close()
    #print(verisontxt)
    t1 = Thread(target=fileserver.start, args=(12345, a))
    t1.start()
    udpt = Thread(target=udpserver)
    udpt.start()

    while True:
        udps.listen()  # 监听连接并等待
        clientsocket, clientinfo = udps.accept()  # 接收连接后获取客户端信息
        # 开启线程处理当前客户端请求
        t = Thread(target=readmsg, args=(clientsocket, clientinfo))
        t.start()
        threadnum = threadnum + 1

        #  t2 = Thread(target=sendnewmessgae, args=(clientsocket, clientinfo))
        # threadnum = threadnum + 1

        # t2.start()
        print("开启的线程数：" + str(threadnum))


main()