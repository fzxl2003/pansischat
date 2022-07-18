# coding=utf-8
# !/usr/bin/python
# -*- coding: UTF-8 -*-
import hashlib
import os
import random
import sys
import threading
import tkinter as tk
import tkinter.filedialog
from tkinter import FLAT, END
from socket import *
from threading import Thread
from cryptography.fernet import Fernet
import webbrowser
from openpyxl import load_workbook, Workbook
import time
import win32api
import win32con
from windnd import windnd
import fileclient
import update
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.asymmetric import padding, rsa
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives import serialization
tcps=socket(AF_INET,SOCK_STREAM)   #创建套接字
version=2000

iptxt=open('组件/ip.txt', mode='r')
ip1=iptxt.read()
iptxt.close()
print(ip1)
ip=(ip1,8080)    #服务器端ip
tcpsituation1=0          #检测网络状况的变量
logcode=''
shezhi_on=1
nowfriendid=''
window=tk.Tk()
window.overrideredirect(True)
window.title('登录页面')
window.geometry('510x258+400+200')
codeshowvariable=tk.StringVar()    #gui窗口创建
indexbgimage= tk.PhotoImage(file="组件/ui/1/登录背景1.png")
offimage0=tk.PhotoImage(file="组件/ui/1/关闭0.png")
minimage0=tk.PhotoImage(file="组件/ui/1/最小化0.png")
offimage1=tk.PhotoImage(file="组件/ui/1/关闭1.png")
minimage1=tk.PhotoImage(file="组件/ui/1/最小化1.png")
uiid = '1'
uimain = '1'
friendlistclick_on = 1
filelistnonum=0
indexbgi=tk.Label(window,image=indexbgimage)
indexbgi.place(x=-5,y=-5)
global usernicheng

userinput=tk.Entry(window,bg='#555555',relief=FLAT,width=25,text='Pansis账户',fg='#8E8E84')
codeinput=tk.Entry(window,bg='#555555',relief=FLAT,width=25,show='',fg='#8E8E84')
tcpsituation=tk.Label(window,text='正在连接服务器',fg='#8E8E84',bg='black')
denglusituation=tk.Label(window,text='',fg='#FF0000',bg='black')
result=tk.Text(window,width=60,height=30,bg='#2B2B2B',fg='#FFFFFF',bd=0)
sendenter=tk.Label(window,width=7,height=1,bg='#535353',bd=0,text='发送')
p2penter = tk.Label(window, width=7, height=1, bg='#535353', bd=0,text='P2P')
p2plabel= tk.Label(window, width=19, height=1, bg='#535353', bd=0,text='当前状态：服务器中转')
windowoff=tk.Label(window,image=offimage0,bg='black')
windowmin=tk.Label(window,image=minimage0,bg='black')
setput=tk.Text(window,width=60,height=4,bg='#535353',bd=0)
window.attributes("-alpha", 0.85)
userinput.place(x=300,y=50)
codeinput.place(x=300,y=85)
tcpsituation.place(x=310,y=115)
windowoff.place(x=465,y=2)
windowmin.place(x=435,y=2)
datatime=time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
nowfile=''
p2psituation=0




def tcpsconnect():
    global tcpsituation1
    global tcps
    global client_private_key,client_public_key,client_public_key1,client_private_key1,client_private_key2
    tcps = socket(AF_INET, SOCK_STREAM)
    try:
        tcps.connect(ip)
    except:
        a=1
        print('连接失败')
    else:
        client_private_key = rsa.generate_private_key(
            public_exponent=65537,
            key_size=2048,
            backend=default_backend()
        )
        client_public_key = client_private_key.public_key()
        # store private key
        client_private_key1 = client_private_key.private_bytes(
            encoding=serialization.Encoding.PEM,
            format=serialization.PrivateFormat.PKCS8,
            encryption_algorithm=serialization.NoEncryption()
        )
        client_public_key1 = client_public_key.public_bytes(
            encoding=serialization.Encoding.PEM,
            format=serialization.PublicFormat.SubjectPublicKeyInfo
        )
        client_private_key2 = serialization.load_pem_private_key(
            client_private_key1,
            password=None,
            backend=default_backend()
        )
        tcps.send(bytes('0024',encoding='utf-8')+client_public_key1)


tcpsconnect()

def tcpsitiontest():
 while True:
  if (update.nowsituation == 0):
     global tcpsituation1
     if tcpsituation1==0:
         tcps.close()
         tcpsconnect()
         tcpsituation.config(text='连接服务器失败')
         tcpsituation1 = 0
     else:


      try:
            tcps.send(('0000'+str(version)).encode('utf-8'))
      except:
            tcps.close()
            tcpsconnect()
            tcpsituation.config(text='连接服务器失败')
            tcpsituation1=0
      else:
            tcpsituation.config(text='连接服务器成功')
            tcpsituation1 = 1
     time.sleep(3)
  else:
      break


def encrypt1(data,cipher_key):     # 进行加密
    data = bytes(data, encoding='utf-8')
    encrypted_text = Fernet(cipher_key).encrypt(data)
    encrypted_text = str(encrypted_text, encoding='utf-8')
    return encrypted_text

def decrypt1(data, cipher_key):  # 进行解密
    if data[:2]=='b\'':              #判定data是否含有b‘   ’
        data=data[2:]                #去掉data中的b‘
        data=data[:-1]               #去掉data中的’
        #print(data)
    data = bytes(data, encoding='utf-8')  #将data由string转为bytes类型
    #print(data)
    #print(cipher_key)
    decrypted_text='1234'
    try:
     decrypted_text = Fernet(cipher_key).decrypt(data)
     decrypted_text = str(decrypted_text, encoding='utf-8')
    except:
        a=1

     #print(decrypted_text)
    return decrypted_text

def codeshow1():            #控制密码框显示
    if codeshowvariable.get()=='0':
        codeinput.config(show='*')
    else:codeinput.config(show='')
    return
def newsend(event):
  global tcpsituation1,p2psituation
  if friendlistclick_on==1:
    if (p2psituation!=1 and p2psituation!=2and p2psituation!=3and p2psituation!=4):
     global filelistnonum
     datatime=time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
     data=setput.get('1.0', 'end')
     result.config(state='normal')
     if len(data)>=1000:
         result.insert('end', '消息过长\n')
         result.see(tk.END)
         result.config(state='disabled')
         return
     i=0
     while i<=len(data)-1:
         if data[i]=='￥' or data[i]=='&' or data[i]=='*' or data[i]=='=' :
             result.insert('end', '不能含有￥&*=的特殊符号\n')
             result.see(tk.END)
             result.config(state='disabled')
             return

         i=i+1
     result.insert('end', usernicheng + '(我)   ' + datatime + '\n' + data + '\n')
     filelistnonum=filelistnonum+1
     result.config(state='disabled')
     data2=datatime+'&'+data
     data2=encrypt1(data2,code)
     data1='0011'+nowfriendid+'&'+userid+'&'+str(data2)
     data1 = encryptplus(data1)
     #print(data1)
     try:
      tcps.send(data1)
     except:
        #tcpsitiontest()
        result.config(state='normal')
        result.insert('end', '网络连接失败\n')
        tcpsituation1 = 0
        tcpsconnect()
     else:
       setput.delete('1.0', 'end')
       for col in ws.iter_cols(min_row=2, min_col=1, max_row=9999, max_col=1):
         for cell in col:
             if cell.value==nowfriendid:
                 cellrow=cell.row
                 break
       for row in ws.iter_rows(min_row=cellrow, min_col=11, max_row=cellrow, max_col=9999):  # 括号代表遍历第一行到第二行,第二列到第三列
         for cell in row:
             if cell.value==None:
                 data=datatime+'&'+userid+data
                 data=encrypt1(data,code)
                 #print('而我却'+data)
                 cell.value=data
                 wb.save(userid+'.xlsx')
                 break
         setput.focus_force()
         win32api.keybd_event(38, 0, 0, 0)
         win32api.keybd_event(38, 0, win32con.KEYEVENTF_KEYUP, 0)
         result.see(tk.END)
         result.config(state='disabled')
    else:
      datatime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
      data = setput.get('1.0', 'end')
      result.config(state='normal')
      if len(data) >= 1000:
          result.insert('end', '消息过长\n')
          result.see(tk.END)
          result.config(state='disabled')
          return
      i = 0
      while i <= len(data) - 1:
          if data[i] == '￥' or data[i] == '&' or data[i] == '*' or data[i] == '=':
              result.insert('end', '不能含有￥&*=的特殊符号\n')
              result.see(tk.END)
              result.config(state='disabled')
              return

          i = i + 1
      result.insert('end', usernicheng + '(我)p2p   ' + datatime + '\n' + data + '\n')
      filelistnonum = filelistnonum + 1
      result.config(state='disabled')
      data1 = '0029' + str(data)
      data1 = p2pencryptplus(data1,p2pkey)
      #print(data1)
      #print(p2padr)

      try:
          udpc.sendto(data1,p2padr)
      except:
          result.config(state='normal')
          result.insert('end', 'p2p连接失败,已恢复服务器中转\n')
          p2psituation = 0
      else:
              setput.delete('1.0', 'end')
              setput.focus_force()
              win32api.keybd_event(38, 0, 0, 0)
              win32api.keybd_event(38, 0, win32con.KEYEVENTF_KEYUP, 0)
              result.see(tk.END)
              result.config(state='disabled')

  return


def mousemove(event):
    global friendlistmainy
    global friendlistmain
    global friendlistmainsum
    if event.delta > 1 and friendlistmainsum!=1:
        friendlistmainy=friendlistmainy+10
    if event.delta < 1 and friendlistmainsum!=len(friendline)-7:
        friendlistmainy=friendlistmainy-10
    #print(friendlistmainy)
    if friendlistmainy<= friendlistmainymin-40:
        friendlistmainsum=friendlistmainsum+1
        friendlistmain=friendlistmain+1
        friendlistmainy=friendlistmainy+45
        if friendlistmain==9:
            friendlistmain=1
        friendlistmoveup()
            #print(friendlistmain)

    if friendlistmainy>=friendlistmainymin+40:
        friendlistmainsum = friendlistmainsum - 1
        friendlistmain=friendlistmain-1
        friendlistmainy=friendlistmainy-45

        if friendlistmain==0:
            friendlistmain=8
        friendlistmovedown()
    friendlistmove()


def friendlistmovedown():
    global friendlistmain
    global friendlistmainsum
    global friendlist8id
    global friendlist1id
    global friendlist2id
    global friendlist3id
    global friendlist4id
    global friendlist5id
    global friendlist6id
    global friendlist7id
    global friendline

    if friendlistmain == 1:
        friendlist1id = friendlistmainsum

        friendlist1.config(text=str(friendline[friendlist1id-1]))
        #print(friendlist1id)
    if friendlistmain == 2:
        friendlist2id = friendlistmainsum

        friendlist2.config(text=str(friendline[friendlist2id-1]))
        #print(friendlist2id)
    if friendlistmain == 3:
        friendlist3id = friendlistmainsum

        friendlist3.config(text=str(friendline[friendlist3id-1]))
        #print(friendlist3id)
    if friendlistmain == 4:
        friendlist4id = friendlistmainsum

        friendlist4.config(text=str(friendline[friendlist4id-1]))
        #print(friendlist4id)
    if friendlistmain == 5:
        friendlist5id = friendlistmainsum

        friendlist5.config(text=str(friendline[friendlist5id-1]))
        #print(friendlist5id)
    if friendlistmain == 6:
        friendlist6id = friendlistmainsum
        friendlist6.config(text=str(friendline[friendlist6id-1]))
        #print(friendlist6id)
    if friendlistmain == 7:
        friendlist7id = friendlistmainsum
        friendlist7.config(text=str(friendline[friendlist7id-1]))
        #print(friendlist7id)
    if friendlistmain == 8:
        friendlist8id = friendlistmainsum
        friendlist8.config(text=str(friendline[friendlist8id-1]))
        #print(friendlist8id)
    friendlistsituation(friendlistmain)

def friendlistmoveup():
    global friendlistmain
    global friendlistmainsum
    global friendlist8id
    global friendlist1id
    global friendlist2id
    global friendlist3id
    global friendlist4id
    global friendlist5id
    global friendlist6id
    global friendlist7id
    global friendline


    if friendlistmain == 1:
        friendlist8id = friendlistmainsum + 7
        friendlist8.config(text=str(friendline[friendlist8id-1]))
        friendlistsituation(8)

        #print(friendlist8id)
    if friendlistmain == 2:
        friendlist1id = friendlistmainsum + 7
        friendlist1.config(text=str(friendline[friendlist1id-1]))
        friendlistsituation(1)
        #print(friendlist1id)
    if friendlistmain == 3:
            friendlist2id = friendlistmainsum + 7
            friendlistsituation(2)
            friendlist2.config(text=str(friendline[friendlist2id - 1]))

         #   print(friendlist2id)
    if friendlistmain == 4:
        friendlist3id = friendlistmainsum + 7
        friendlistsituation(3)
        friendlist3.config(text=str(friendline[friendlist3id-1]))
        #print(friendlist3id)
    if friendlistmain == 5:
        friendlist4id = friendlistmainsum + 7
        friendlistsituation(4)
        friendlist4.config(text=str(friendline[friendlist4id-1]))
        #print(friendlist4id)
    if friendlistmain == 6:
        friendlist5id = friendlistmainsum + 7
        friendlistsituation(5)
        friendlist5.config(text=str(friendline[friendlist5id-1]))
        #print(friendlist5id)
    if friendlistmain == 7:
        friendlist6id = friendlistmainsum + 7
        friendlistsituation(6)
        friendlist6.config(text=str(friendline[friendlist6id-1]))
        #print(friendlist6id)
    if friendlistmain == 8:
        #print('friendlistmainsum' + str(friendlistmainsum))
        friendlistsituation(7)
        friendlist7id = friendlistmainsum + 7
        friendlist7.config(text=str(friendline[friendlist7id-1]))





def friendlistmove():   #好友列表滚动
    if friendlistmain==1:
        friendlist1.place(x=friendlistmainx, y=friendlistmainy)
        friendlist2.place(x=friendlistmainx, y=friendlistmainy+45*1)
        friendlist3.place(x=friendlistmainx, y=friendlistmainy + 45*2)
        friendlist4.place(x=friendlistmainx, y=friendlistmainy + 45*3)
        friendlist5.place(x=friendlistmainx, y=friendlistmainy + 45*4)
        friendlist6.place(x=friendlistmainx, y=friendlistmainy + 45*5)
        friendlist7.place(x=friendlistmainx, y=friendlistmainy + 45*6)
        friendlist8.place(x=friendlistmainx, y=friendlistmainy + 45*7)

    if friendlistmain==2:
        friendlist2.place(x=friendlistmainx, y=friendlistmainy)
        friendlist3.place(x=friendlistmainx, y=friendlistmainy+45*1)
        friendlist4.place(x=friendlistmainx, y=friendlistmainy + 45*2)
        friendlist5.place(x=friendlistmainx, y=friendlistmainy + 45*3)
        friendlist6.place(x=friendlistmainx, y=friendlistmainy + 45*4)
        friendlist7.place(x=friendlistmainx, y=friendlistmainy + 45*5)
        friendlist8.place(x=friendlistmainx, y=friendlistmainy + 45*6)
        friendlist1.place(x=friendlistmainx, y=friendlistmainy + 45*7)
    if friendlistmain==3:
        friendlist3.place(x=friendlistmainx, y=friendlistmainy)
        friendlist4.place(x=friendlistmainx, y=friendlistmainy+45*1)
        friendlist5.place(x=friendlistmainx, y=friendlistmainy + 45*2)
        friendlist6.place(x=friendlistmainx, y=friendlistmainy + 45*3)
        friendlist7.place(x=friendlistmainx, y=friendlistmainy + 45*4)
        friendlist8.place(x=friendlistmainx, y=friendlistmainy + 45*5)
        friendlist1.place(x=friendlistmainx, y=friendlistmainy + 45*6)
        friendlist2.place(x=friendlistmainx, y=friendlistmainy + 45*7)
    if friendlistmain==4:
        friendlist4.place(x=friendlistmainx, y=friendlistmainy)
        friendlist5.place(x=friendlistmainx, y=friendlistmainy+45*1)
        friendlist6.place(x=friendlistmainx, y=friendlistmainy + 45*2)
        friendlist7.place(x=friendlistmainx, y=friendlistmainy + 45*3)
        friendlist8.place(x=friendlistmainx, y=friendlistmainy + 45*4)
        friendlist1.place(x=friendlistmainx, y=friendlistmainy + 45*5)
        friendlist2.place(x=friendlistmainx, y=friendlistmainy + 45*6)
        friendlist3.place(x=friendlistmainx, y=friendlistmainy + 45*7)
    if friendlistmain==5:
        friendlist5.place(x=friendlistmainx, y=friendlistmainy)
        friendlist6.place(x=friendlistmainx, y=friendlistmainy+45*1)
        friendlist7.place(x=friendlistmainx, y=friendlistmainy + 45*2)
        friendlist8.place(x=friendlistmainx, y=friendlistmainy + 45*3)
        friendlist1.place(x=friendlistmainx, y=friendlistmainy + 45*4)
        friendlist2.place(x=friendlistmainx, y=friendlistmainy + 45*5)
        friendlist3.place(x=friendlistmainx, y=friendlistmainy + 45*6)
        friendlist4.place(x=friendlistmainx, y=friendlistmainy + 45*7)
    if friendlistmain==6:
        friendlist6.place(x=friendlistmainx, y=friendlistmainy)
        friendlist7.place(x=friendlistmainx, y=friendlistmainy+45*1)
        friendlist8.place(x=friendlistmainx, y=friendlistmainy + 45*2)
        friendlist1.place(x=friendlistmainx, y=friendlistmainy + 45*3)
        friendlist2.place(x=friendlistmainx, y=friendlistmainy + 45*4)
        friendlist3.place(x=friendlistmainx, y=friendlistmainy + 45*5)
        friendlist4.place(x=friendlistmainx, y=friendlistmainy + 45*6)
        friendlist5.place(x=friendlistmainx, y=friendlistmainy + 45*7)
    if friendlistmain==7:
        friendlist7.place(x=friendlistmainx, y=friendlistmainy)
        friendlist8.place(x=friendlistmainx, y=friendlistmainy+45*1)
        friendlist1.place(x=friendlistmainx, y=friendlistmainy + 45*2)
        friendlist2.place(x=friendlistmainx, y=friendlistmainy + 45*3)
        friendlist3.place(x=friendlistmainx, y=friendlistmainy + 45*4)
        friendlist4.place(x=friendlistmainx, y=friendlistmainy + 45*5)
        friendlist5.place(x=friendlistmainx, y=friendlistmainy + 45*6)
        friendlist6.place(x=friendlistmainx, y=friendlistmainy + 45*7)
    if friendlistmain==8:
        friendlist8.place(x=friendlistmainx, y=friendlistmainy)
        friendlist1.place(x=friendlistmainx, y=friendlistmainy+45*1)
        friendlist2.place(x=friendlistmainx, y=friendlistmainy + 45*2)
        friendlist3.place(x=friendlistmainx, y=friendlistmainy + 45*3)
        friendlist4.place(x=friendlistmainx, y=friendlistmainy + 45*4)
        friendlist5.place(x=friendlistmainx, y=friendlistmainy + 45*5)
        friendlist6.place(x=friendlistmainx, y=friendlistmainy + 45*6)
        friendlist7.place(x=friendlistmainx, y=friendlistmainy + 45*7)



def friendlistsituation(friendlistnum):
    global friendlistmain
    global friendlistmainsum
    global friendlist8id
    global friendlist1id
    global friendlist2id
    global friendlist3id
    global friendlist4id
    global friendlist5id
    global friendlist6id
    global friendlist7id
    global friendline
    #print(len(friendline))
    if friendlistnum==1 or friendlistnum==0 and len(friendline)>=1:
        for col in ws.iter_cols(min_row=2, min_col=1, max_col=1, max_row=999):
            for cell in col:
                if cell.value==friendline[friendlist1id-1]:
                    if str(ws["D" + str(cell.row)].value) == '1':
                        friendlist1.config(fg='red')
                    else:friendlist1.config(fg='#8E8E84')

    if friendlistnum==2 or friendlistnum==0 and len(friendline)>=2:
        for col in ws.iter_cols(min_row=2, min_col=1, max_col=1, max_row=999):
            for cell in col:
                if cell.value==friendline[friendlist2id-1]:
                    if str(ws["D" + str(cell.row)].value) == '1':
                        friendlist2.config(fg='red')
                    else:friendlist2.config(fg='#8E8E84')
    if friendlistnum==3 or friendlistnum==0 and len(friendline)>=3:
        for col in ws.iter_cols(min_row=2, min_col=1, max_col=1, max_row=999):
            for cell in col:
                if cell.value==friendline[friendlist3id-1]:
                    if str(ws["D" + str(cell.row)].value) == '1':
                        friendlist3.config(fg='red')
                    else:friendlist3.config(fg='#8E8E84')
    if friendlistnum==4 or friendlistnum==0 and len(friendline)>=4:
        for col in ws.iter_cols(min_row=2, min_col=1, max_col=1, max_row=999):
            for cell in col:
                if cell.value==friendline[friendlist4id-1]:
                    if str(ws["D" + str(cell.row)].value) == '1':
                        friendlist4.config(fg='red')
                    else:friendlist4.config(fg='#8E8E84')
    if friendlistnum==5 or friendlistnum==0 and len(friendline)>=5:
        for col in ws.iter_cols(min_row=2, min_col=1, max_col=1, max_row=999):
            for cell in col:
                if cell.value==friendline[friendlist5id-1]:
                    if str(ws["D" + str(cell.row)].value) == '1':
                        friendlist5.config(fg='red')
                    else:friendlist5.config(fg='#8E8E84')
    if friendlistnum==6 or friendlistnum==0 and len(friendline)>=6:
        for col in ws.iter_cols(min_row=2, min_col=1, max_col=1, max_row=999):
            for cell in col:
                if cell.value==friendline[friendlist6id-1]:
                    if str(ws["D" + str(cell.row)].value) == '1':
                        friendlist6.config(fg='red')
                    else:friendlist6.config(fg='#8E8E84')
    if friendlistnum==7 or friendlistnum==0 and len(friendline)>=7:
        for col in ws.iter_cols(min_row=2, min_col=1, max_col=1, max_row=999):
            for cell in col:
                if cell.value==friendline[friendlist7id-1]:
                    if str(ws["D" + str(cell.row)].value) == '1':
                        friendlist7.config(fg='red')
                    else:friendlist7.config(fg='#8E8E84')


def friendlist1click(event):
    global nowfriendid,nowfriendlistclickid
 #   print('1')
    if friendlistclick_on==1 and nowfriendid!=friendline[friendlist1id - 1]:
     nowfriendid=friendline[friendlist1id-1]
     nowfriendlistclickid=1
     friendlistclick()


def friendlist2click(event):
    global nowfriendid,nowfriendlistclickid
    if friendlistclick_on == 1 and nowfriendid!=friendline[friendlist2id - 1]:
      nowfriendid = friendline[friendlist2id - 1]
      nowfriendlistclickid = 2
      friendlistclick()
def friendlist3click(event):
    global nowfriendid,nowfriendlistclickid
    if friendlistclick_on == 1 and nowfriendid!=friendline[friendlist3id - 1]:
      nowfriendid = friendline[friendlist3id - 1]
      nowfriendlistclickid = 3
      friendlistclick()
def friendlist4click(event):
    global nowfriendid,nowfriendlistclickid
    if friendlistclick_on == 1 and nowfriendid!=friendline[friendlist4id - 1]:
      nowfriendid = friendline[friendlist4id - 1]
      nowfriendlistclickid = 4
      friendlistclick()
def friendlist5click(event):
    global nowfriendid,nowfriendlistclickid
    if friendlistclick_on == 1 and nowfriendid!=friendline[friendlist5id - 1]:
      nowfriendid = friendline[friendlist5id - 1]
      nowfriendlistclickid = 5
      friendlistclick()
def friendlist6click(event):
    global nowfriendid,nowfriendlistclickid
    if friendlistclick_on == 1 and nowfriendid!=friendline[friendlist6id - 1]:
      nowfriendid = friendline[friendlist6id - 1]
      nowfriendlistclickid = 6
      friendlistclick()
def friendlist7click(event):
    global nowfriendid,nowfriendlistclickid
    if friendlistclick_on == 1 and nowfriendid!=friendline[friendlist7id - 1]:
      nowfriendid = friendline[friendlist7id - 1]
      nowfriendlistclickid = 7
      friendlistclick()
def friendlist8click(event):
    global nowfriendid,nowfriendlistclickid
    if friendlistclick_on == 1 and nowfriendid!=friendline[friendlist8id - 1]:
      nowfriendid = friendline[friendlist8id - 1]
      nowfriendlistclickid = 8
      friendlistclick()

def friendlistclick():

    global friendlistclick_on
    #print('1')
    result.config(state='disabled')
    friendlistclick_on = 0
    p2psituationcontrol(0)
    global usernicheng
    global friendnicheng
    global nowfriendrow,sendenter,p2penter

    result.place(x=350, y=45)
    setput.place(x=350, y=440)

    sendenter.place(x=718, y=497)

    p2penter.place(x=490, y=497)
    p2plabel.place(x=350, y=497)
    for col in ws.iter_cols(min_row=2, min_col=1, max_col=1,max_row=999):
     for cell in col:
        if cell.value==nowfriendid:
            nowfriendrow=cell.row
            usernicheng=ws['B1'].value
            break
    result.config(state='normal')
    result.delete('1.0', 'end')
    tcp('0013'+nowfriendid)

def fileopen(filename):
    m=filename
    if os.path.exists(filelocation[filename]+'/'+filename)!=True:
      if filename[0] == '(':
        i = 1
        while i < len(filename):
            if filename[i] == ')':
                filename = filename[i + 1:]
                break
            i = i + 1
        if os.path.exists(filelocation[m] + '/' + filename) != True:
            result.config(state='normal')
            result.delete(str(filelistline[m]) + '.0', str(filelistline[m]) + '.end')
            result.insert(str(filelistline[m]) + '.0', m + "文件不存在(点击重新下载)")
            result.config(state='disabled')
            downloadaddfiletag(m, filelistline[m])
            for row in ws.iter_rows(min_row=nowfriendrow, min_col=filelistno[m], max_row=nowfriendrow,
                                    max_col=filelistno[m]):  # 括号代表遍历第一行到第二行,第二列到第三列
                for cell in row:
                    if cell.value != None:
                        data = decrypt1(cell.value, code)
                        i = 1
                        while i <= len(data):
                            if data[len(data) - i] == '&':
                                data = data[:-i + 1]
                                break
                            i = i + 1
                        data = data + "*"
                        cell.value = encrypt1(data, code)
                        wb.save(userid + ".xlsx")
                        filelocation[m] = "*"
        else:
            t3=Thread(target=openfile, args=(filelocation[m] + '/' + filename, 'nowfilelocation'))
            t3.start()


    else:
        t3 = Thread(target=openfile, args=(filelocation[m] + '/' + filename, 'nowfilelocation'))
        t3.start()
    return
def openfile(location,f):


    b = os.path.join(location)
    #print("qwds11" + b)
    os.startfile(location)
   # os.system(f'start {os.path.realpath(location)}')
    #print("qwds" + location)

def fileopenfold(filename):
    webbrowser.open(filelocation[filename], new=2)
    return
def filedelete(filename):
    global friendlistclick_on
    try:
     os.remove(filelocation[filename]+'/'+filename)
    except:
        a=1
    else:
        a=1
    fileclient.filesituation = -2
    result.config(state='normal')
    result.delete(str(filelistline[filename]) + '.0', str(filelistline[filename]) + '.end')
    result.insert(str(filelistline[filename]) + '.0', filename + "(未下载)  点击下载")
    downloadaddfiletag(filename, filelistline[filename])
    result.config(state='disabled')
    friendlistclick_on=1
    for row in ws.iter_rows(min_row=nowfriendrow, min_col=filelistno[filename], max_row=nowfriendrow,
                            max_col=filelistno[filename]):  # 括号代表遍历第一行到第二行,第二列到第三列
        for cell in row:
            if cell.value != None:
                data = decrypt1(cell.value, code)
                i = 1
                while i <= len(data):
                    if data[len(data) - i] == '&':
                        data = data[:-i + 1]
                        break
                    i = i + 1
                data = data + "*"
                cell.value = encrypt1(data, code)
                wb.save(userid + ".xlsx")
                filelocation[filename] = "*"
    return
def downloadaddfiletag(filename,resultlinenum):       #未下载文件的标签加入
    result.config(state='normal')
    result.tag_add(filename, str(resultlinenum) + '.0', str(resultlinenum) + '.end')
    result.tag_config(filename, foreground='red', font='黑体 12')  # 再为tag1标签进行设置
    result.tag_bind(filename, "<Enter>",
                    lambda event: result.config(cursor='hand2'))  # 鼠标移入,鼠标样式变手式
    result.tag_bind(filename, "<Leave>",
                    lambda event: result.config(cursor='xterm'))  # 鼠标离开,鼠标样式变 I
    result.tag_bind(filename, "<Button-1>", lambda event: download(filename))
    result.config(state='disabled')
    return
def downloadedaddfiletag(filename):      #已下载文件的标签加入
    resultlinetxt = result.get(str(filelistline[filename]) + '.0', str(filelistline[filename]) + '.end')
    a = 0
    typenum = 1
    while a <= len(resultlinetxt) - 1:
        if resultlinetxt[a] == "打" or resultlinetxt[a] == "删":
            if typenum == 1:
                start1 = a
            if typenum == 2:
                start2 = a
            if typenum == 3:
                start3 = a
                break
            typenum = typenum + 1
        a = a + 1
    result.config(state='normal')
    result.tag_add(filename + 'open', str(filelistline[filename]) + '.' + str(start1),
                   str(filelistline[filename]) + '.' + str(start1 + 2))
    result.tag_config(filename+ 'open', foreground='red', font='黑体 12')  # 再为tag1标签进行设置
    result.tag_bind(filename + 'open', "<Enter>",
                    lambda event: result.config(cursor='hand2'))  # 鼠标移入,鼠标样式变手式
    result.tag_bind(filename + 'open', "<Leave>",
                    lambda event: result.config(cursor='xterm'))  # 鼠标离开,鼠标样式变 I
    result.tag_bind(filename + 'open', "<Button-1>", lambda event: fileopen(filename))

    result.tag_add(filename + 'openfold', str(filelistline[filename]) + '.' + str(start2),
                   str(filelistline[filename]) + '.' + str(start2 + 5))
    result.tag_config(filename+ 'openfold', foreground='red', font='黑体 12')  # 再为tag1标签进行设置
    result.tag_bind(filename + 'openfold', "<Enter>",
                    lambda event: result.config(cursor='hand2'))  # 鼠标移入,鼠标样式变手式
    result.tag_bind(filename + 'openfold', "<Leave>",
                    lambda event: result.config(cursor='xterm'))  # 鼠标离开,鼠标样式变 I
    result.tag_bind(filename + 'openfold', "<Button-1>", lambda event: fileopenfold(filename))

    result.tag_add(filename + 'delete', str(filelistline[filename]) + '.' + str(start3),
                   str(filelistline[filename]) + '.' + str(start3 + 4))
    result.tag_config(filename+ 'delete', foreground='red', font='黑体 12')  # 再为tag1标签进行设置
    result.tag_bind(filename + 'delete', "<Enter>",
                    lambda event: result.config(cursor='hand2'))  # 鼠标移入,鼠标样式变手式
    result.tag_bind(filename + 'delete', "<Leave>",
                    lambda event: result.config(cursor='xterm'))  # 鼠标离开,鼠标样式变 I
    result.tag_bind(filename + 'delete', "<Button-1>", lambda event: filedelete(filename))
    result.config(state='disabled')
    return
def download1(filename,nowfilelocation):
    global friendlistclick_on
    fileclient.filesituation=1
    friendlistclick_on = 0
    fileclient.download1(filename,nowfilelocation)
    return
def download2(filename,nowfilelocation):
    global friendlistclick_on
    timeout=0
    while True:
        try:
            #print("123:"+fileclient.filesituationtxt)
            time.sleep(1)
        except:
            a=1
        else:
            result.config(state='normal')
            result.delete(str(filelistline[filename]) + '.0', str(filelistline[filename]) + '.end')
            result.insert(str(filelistline[filename]) + '.0', filename + fileclient.filesituationtxt+'   取消下载')
            result.tag_add(filename + 'cancel', str(filelistline[filename]) + '.' + str(len(filename + fileclient.filesituationtxt)+3),
                           str(filelistline[filename]) + '.end')
            result.tag_config(filename+ 'cancel', foreground='red', font='黑体 12')  # 再为tag1标签进行设置
            result.tag_bind(filename + 'cancel', "<Enter>",
                            lambda event: result.config(cursor='hand2'))  # 鼠标移入,鼠标样式变手式
            result.tag_bind(filename + 'cancel', "<Leave>",
                            lambda event: result.config(cursor='xterm'))  # 鼠标离开,鼠标样式变 I
            result.tag_bind(filename + 'cancel', "<Button-1>", lambda event: filedelete(filename))
            if fileclient.filesituationtxt[0:4]=="下载失败":
                timeout=timeout+1
                if timeout==3:
                    result.delete(str(filelistline[filename]) + '.0', str(filelistline[filename]) + '.end')
                    result.insert(str(filelistline[filename]) + '.end', filename+"  点击重新下载")
                    downloadaddfiletag(filename, filelistline[filename])
                    result.config(state='disabled')
                    break
            if fileclient.filesituation==-1:
                result.delete(str(filelistline[filename]) + '.0', str(filelistline[filename]) + '.end')
                result.insert(str(filelistline[filename]) + '.end', filename+"(已下载)  打开  打开文件夹  删除文件")
                downloadedaddfiletag(filename)
                for row in ws.iter_rows(min_row=nowfriendrow, min_col=filelistno[filename], max_row=nowfriendrow,
                                        max_col=filelistno[filename]):  # 括号代表遍历第一行到第二行,第二列到第三列
                    for cell in row:
                        if cell.value != None:
                            data = decrypt1(cell.value, code)
                            data=data[:-1]
                            data=data+nowfilelocation
                            filelocation[filename]=nowfilelocation
                            cell.value=encrypt1(data,code)
                            wb.save(userid + ".xlsx")
                friendlistclick_on=1
                fileclient.filesituation = 0
                break
            result.config(state='disabled')


def download(filename):
    global nowfilelocation
    global nowfilename
    if filelocation[filename]=="*" :
        nowfilelocation=tkinter.filedialog.askdirectory()
        if nowfilelocation!='':
         #print(nowfilelocation)
         data="0020"
         tcps.send(data.encode('utf-8'))
         nowfilename=filename
         t1 = Thread(target=download1, args=(nowfilename, nowfilelocation))
         t2 = Thread(target=download2, args=(nowfilename, nowfilelocation))
         t1.start()
         t2.start()
         result.tag_delete(nowfilename)


    return

def uploaddelete(line):
    fileclient.uploadsituation=-2




def upload(files):
 if friendlistclick_on==1:
   msg = '\n'.join((item.decode('gbk') for item in files))
   msg=str(msg)
   #print('sdhajk:'+msg)
   if os.path.isfile(msg):
    if os.path.getsize(msg)!=0:
     i=1
     while i <= len(msg):
        if msg[len(msg) - i] == '\\':
            filename = msg[len(msg) - i + 1:]
            break
        i = i + 1
     i = 1
     msg1=msg
     while i <= len(msg1) - 1:
         if msg1[i] == "\\":
             msg1 = msg1[:i] + '/' + msg1[i + 1:]
         i = i + 1
     i = 1
     while i <= len(msg1) - 1:
         if msg1[len(msg1) - i] == "/":
             msg1 = msg1[:len(msg1) - i]
         i = i + 1
     filelocation[filename]=msg1
     global filelistnonum
     datatime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
     result.config(state='normal')
     result.insert('end', usernicheng + '(我)   ' + datatime + '\n')
     result.insert('end', usernicheng + '(我)   ' + datatime + '\n')
     result.see(tk.END)
     result.config(state='disabled')
     t1=Thread(target=upload1, args=(filename,msg))
     t1.start()
     t2 = Thread(target=upload2, args=(filename,msg1))
     t2.start()
     return

def uploadwait(a):
    global friendlistclick_on
    time.sleep(10)
    friendlistclick_on = 1

def upload1(filename, msg):
    global friendlistclick_on
    fileclient.uploadfilesituation = 1
    friendlistclick_on = 0
    senddata='0022'+filename
    tcps.send(senddata.encode('utf-8'))

    fileclient.upload1(filename, msg)
    return
def upload2(filename,nowfilelocation):
    global friendlistclick_on
    global newuploadfileline
    timeout=0

    while True:
        filelistline = int(result.index('end-1c').split('.')[0]) - 1
        friendlistclick_on = 0
        try:
            #print("123:"+fileclient.filesituationtxt)
            time.sleep(1)
        except:
            a=1
        else:
            result.config(state='normal')
            result.delete(str(filelistline) + '.0', str(filelistline) + '.end')
            result.insert(str(filelistline) + '.0', filename + fileclient.uploadfilesituationtxt+'   取消上传')
            result.tag_add(filename + 'uploadcancel', str(filelistline) + '.' + str(len(filename + fileclient.uploadfilesituationtxt)+3),
                           str(filelistline) + '.end')
            result.tag_config(filename+ 'uploadcancel', foreground='red', font='黑体 12')  # 再为tag1标签进行设置
            result.tag_bind(filename + 'uploadcancel', "<Enter>",
                            lambda event: result.config(cursor='hand2'))  # 鼠标移入,鼠标样式变手式
            result.tag_bind(filename + 'uploadcancel', "<Leave>",
                            lambda event: result.config(cursor='xterm'))  # 鼠标离开,鼠标样式变 I
            result.tag_bind(filename + 'uploadcancel', "<Button-1>", lambda event: uploaddelete(filelistline))
            if fileclient.uploadfilesituationtxt[0:4]=="上传失败":
                    a=result.get(str(filelistline) + '.0', str(filelistline) + '.end')
                    result.delete(str(filelistline) + '.0', str(filelistline) + '.end')
                    result.insert('end', a)
                    friendlistclick_on = 1
                    fileclient.uploadfilesituation = 0
                    break
            if fileclient.uploadfilesituation==-1:
                result.delete(str(filelistline) + '.0', str(filelistline) + '.end')
                result.insert(str(filelistline) + '.end', filename+"上传完成")
                newuploadfileline=filelistline
                t1 = Thread(target=uploadwait, args=(filename))
                t1.start()

                fileclient.uploadfilesituation = 0
                break
            result.config(state='disabled')
            result.see(tk.END)







filelistline={}    #记录文件所在行数
filelistno={}      #记录文件所在聊天的第几个
filelocation={}    #记录文件路径
def friendlistclick2():
    global usernicheng
    global friendnicheng
    global nowfriendrow
    global filelist
    global filelistline,p2penter,sendenter
    global filelistnonum,friendlistclick_on,nowfriendlistclickid
    filelistline.clear()
    filelistno.clear()
    filelocation.clear()
    resultlinenum=0
    filelistnonum=11
    ws["D" + str(nowfriendrow)].value = '0'
    wb.save(userid + ".xlsx")
    friendlistsituation(nowfriendlistclickid)
    for row in ws.iter_rows(min_row=nowfriendrow, min_col=11, max_row=nowfriendrow,max_col=999):  # 括号代表遍历第一行到第二行,第二列到第三列
        for cell in row:

            if cell.value!=None:
                data=decrypt1(cell.value,code)
                #print("!!!"+data)
                i=0


                while i<=len(data)-1:
                    if data[i]=='&':
                        dataid=data[i+1:i+6]

                        datatime=data[:i]

                        data=data[i+6:]
                        #print(data)

                        if dataid == userid:
                            dataid = usernicheng + '(我)'

                        else:
                            dataid = dataid + '  ' + str(friendnicheng) + '(对方)'
                        if data[:3]=='$$$':

                            a=2
                            while a <= len(data) - 1:
                                if data[a] == '&':
                                    result.config(state='normal')

                                    filelocation[data[3:a]]=data[a+1:]  #  data[0:a]文件名  data[a+1:]文件路径     #！！！！！！！！！！！！！！1切换删除
                                    #print("uuds\n")
                                    #print(filelocation[data[3:a]])
                                    if data[a+1:]=="*":
                                        buttontext=data[3:a] +"(未下载)  点击下载"
                                        result.insert('end', dataid + '   ' + datatime + '\n')
                                        result.insert('end', buttontext + '\n')
                                      #  resultlinenum = resultlinenum + 2
                                        #print(result.index('end-1c').split('.')[0])

                                        filelistline[data[3:a]] = int(result.index('end-1c').split('.')[0])-1  # ！！！！！！！！！！！！！！1切换删除
                                        filelistno[data[3:a]] = filelistnonum  # ！！！！！！！！！！！！！！1切换删除
                                        downloadaddfiletag(data[3:a], filelistline[data[3:a]])
                                    else:
                                        buttontext=data[3:a] +"(已下载)  打开  打开文件夹  删除文件"
                                        result.insert('end', dataid + '   ' + datatime + '\n')
                                        result.insert('end', buttontext + '\n')
                                      #  resultlinenum = resultlinenum + 2
                                        filelistline[data[3:a]] = int(result.index('end-1c').split('.')[0])-1  # ！！！！！！！！！！！！！！1切换删除
                                        filelistno[data[3:a]] = filelistnonum  # ！！！！！！！！！！！！！！1切换删除
                                        downloadedaddfiletag(data[3:a])
                                #    result.tag_add(data[3:a],str(resultlinenum)+'.0',str(resultlinenum)+'.end')
                                 #   result.tag_config(data[3:a], foreground='red',font='黑体 12')  # 再为tag1标签进行设置
                                 #   result.tag_bind(data[3:a], "<Enter>",
                                #                 lambda event: result.config(cursor='hand2'))  # 鼠标移入,鼠标样式变手式
                                 #   result.tag_bind(data[3:a], "<Leave>",
                                 #                lambda event: result.config(cursor='xterm'))  # 鼠标离开,鼠标样式变 I
                                 #   result.tag_bind(data[3:a],"<Button-1>", lambda event:download(data[3:a]))
                                 #
                                 #   result.delete(str(filelistline[data[3:a]])+'.0',str(filelistline[data[3:a]])+'.'+str(len(buttontext)-1))
                                 #   result.insert(str(filelistline[data[3:a]])+'.0',"1234")

                                   # tk.Button(window, height=1,text=buttontext,command=lambda arg=data[3:a]:download(arg)).pack()
                                    break
                                a=a+1
                        else:
                            result.config(state='normal')
                            result.insert('end',dataid+'   '+datatime+'\n'+data+'\n')
                           # resultlinenum = resultlinenum + 2
                        break
                    i=i+1
                filelistnonum=filelistnonum+1
                result.see(tk.END)
    setput.focus_force()
    win32api.keybd_event(38, 0, 0, 0)
    win32api.keybd_event(38, 0, win32con.KEYEVENTF_KEYUP, 0)

    result.see(tk.END)
    result.config(state='disabled')
    friendlistclick_on = 1
    setput.bind('<Return>',newsend)
    sendenter.bind('<Button-1>', newsend)
    p2penter.bind('<Button-1>', p2penter1)


p2psituation=0
def udpwait1():
    i=30
    global p2psituation
    while(i>0):
        time.sleep(1)
        i=i-1
        p2plabel.config(text='正在请求连接'+str(i)+'s')
        if (p2psituation==1 or p2psituation==2 ):
            result.insert('end', 'P2P加密通信连接成功，注意此模式下聊天记录不会被保存,文件传输依旧通过服务器中转，切换聊天好友将自动退出P2P模式\n')
            p2plabel.config(text='当前状态：P2P加密')
            break
        if ( p2psituation==5 or p2psituation==0 ):
            p2psituationcontrol(0)
    if (p2psituation==-1):
        p2psituationcontrol(0)

def p2penter1(event):
    global udpc,p2psituation
    if (p2psituation==0):
        p2plabel.config(text='正在请求连接30s')
        udpc = socket(AF_INET, SOCK_DGRAM)  # 创建套接字
        udpc.bind(('0.0.0.0', random.randint(3456,8765)))
        udpc.sendto(b'0015', (ip1,7583))
        p2psituation=-1
        t1 = threading.Thread(target=udpwait1)
        t1.setDaemon(True)
        t1.start()
        t1 = threading.Thread(target=p2preceive1)
        t1.setDaemon(True)
        t1.start()
    if (p2psituation==1 or p2psituation==2 or p2psituation==3 or p2psituation==4):
        p2psituationcontrol(0)

def p2psituationcontrol(change):    #change为-2代表不改变，仅刷新
    global p2psituation

    if (change!=-2):
        p2psituation=change
    if (p2psituation==3 or p2psituation==4):
        p2plabel.config(text='当前状态：尝试重连中')
    if (p2psituation==1 or p2psituation==2):
        p2plabel.config(text='当前状态：P2P加密')

    if (p2psituation==0 or p2psituation==5):
        try:
           udpc.close()
        except:
            a=1
        p2psituation=0
        p2plabel.config(text='当前状态：服务器中转')

def p2pbegin(data):
    global p2pkey,p2padr,p2psituation
    p2padr1= []
    begin = 0
    i = 1
    strlen = len(data)
    leibie = 1
    while (i < strlen):
        if (data[i] == '&'):
            shuju = data[begin:i ]
            #print(shuju)
            begin = i+1
            leibie = leibie + 1
            if (leibie==2):
                p2padr1.append(shuju)
            if (leibie==3):
                p2padr1.append(int(shuju))

                data = data[begin + 2:-1]
                begin = 0
                newdata = b''
                strlen = len(data)
                i = 1
                while (i < strlen):
                    # print(data[i])
                    if (data[i] == '\\'):
                        newdata = newdata + b'\n' + bytes(data[begin + 1:i ], encoding='utf-8')
                        begin = i + 1
                    i = i + 1
                newdata = b'-'+newdata[1:] + b'\n'
                publickeyy = serialization.load_pem_public_key(newdata, backend=default_backend())
                p2pkey['friendkey'] = publickeyy

        i = i + 1
    p2padr=(p2padr1[0],p2padr1[1])
    p2psituation=1
    t1 = threading.Thread(target=p2pxintiao,args=(p2padr,p2pkey))
    t1.setDaemon(True)
    t1.start()
    t2 = threading.Thread(target=p2preceive, args=(p2padr, p2pkey))
    t2.setDaemon(True)
    t2.start()
def p2preceive1():
    #print('1234')
    #print(p2psituation)
    while (p2psituation == -1):
        data1 = udpc.recv(10240)
        #print('123')
        data1 = data1.decode(encoding='utf-8')
        if (data1[:4] == '0016'):
            ser0026(data1[4:])
            break
def p2pxintiao(p2padr,p2pkey):

    global p2psituation
    #print(p2psituation)
    while (p2psituation!=0 and p2psituation!=-1):
     try:
       udpc.sendto(b'0028',p2padr)
     except:
         p2psituationcontrol(0)
     else:
      print(p2psituation)
      time.sleep(3)
      p2psituation=p2psituation+1
      if (p2psituation==5):
         p2psituationcontrol(0)


def p2preceive(p2padr,p2pkey):
    global p2psituation

    while (p2psituation!=0 and p2psituation!=-1 and p2psituation!=5):
      try:
           data1=udpc.recvfrom(10240)
      except:
            p2psituationcontrol(0)
      else:
        nowadr=data1[1]
        data = data1[0]

        if (nowadr[0]==p2padr[0] and nowadr[1]==p2padr[1]):
            p2psituationcontrol(1)
            if (data!=b'0028'):
                data=p2pdecryptplus(data,p2pkey)
                zhiling=data[:4]
                data=data[4:]
                if (zhiling=='0029'):
                    p2pmessage(data)



def p2pmessage(data):
    dataid = nowfriendid + '  ' + str(friendnicheng) + '(对方)P2P'
    result.config(state='normal')
    result.insert('end', dataid + '   ' + datatime + '\n' + data + '\n')
    result.see(tk.END)



def ser0026(receivedata):
    #print('aaaa')
    global p2pkey,p2psituation,tcpsituation1
    i = 4
    strlen=len(receivedata)
    #print(receivedata)
    while (i<strlen):
       if (receivedata[i] == '&'):
           begin = i
           break
       i = i + 1
    gip = receivedata[:i]
    gport = int(receivedata[i + 1:])
    #print(gport)
    p2pprivate_key = rsa.generate_private_key(
        public_exponent=65537,
        key_size=2048,
        backend=default_backend()
    )
    p2ppublic_key = p2pprivate_key.public_key()
    # store private key
    p2pprivate_key1 = p2pprivate_key.private_bytes(
        encoding=serialization.Encoding.PEM,
        format=serialization.PrivateFormat.PKCS8,
        encryption_algorithm=serialization.NoEncryption()
    )
    p2ppublic_key1 = p2ppublic_key.public_bytes(
        encoding=serialization.Encoding.PEM,
        format=serialization.PublicFormat.SubjectPublicKeyInfo
    )
    p2pprivate_key2 = serialization.load_pem_private_key(
        p2pprivate_key1,
        password=None,
        backend=default_backend()
    )

    data1 = '0026' + userid + nowfriendid + gip + '&' + str(gport) + '&' + str(p2ppublic_key1)
    #print(data1)
    data1 = encryptplus(data1)
    #print('data1'+str(data1))
    try:
        tcps.send(data1)
    except:
        #print('hjfdks')
        result.config(state='normal')
        result.insert('end', '网络连接失败\n')
        tcpsituation1 = 0
        p2psituation=0
        p2plabel.config(text='当前状态：服务器中转')
        tcpsconnect()
    p2pkey={}
    #print('data1')
    p2pkey['private_key']=p2pprivate_key2
    p2pkey['public_key'] = p2ppublic_key1


def p2pdecryptplus(receivedata,p2pkey):
    global tcpsituation1,p2psituation
    #print(b'abcdeabcf'+receivedata)

    try:
        original_message = b''
        if receivedata[:3] == b'&&&':
            i = 3
            beginnum = 3
            while i < len(receivedata):


                if receivedata[i:i + 3] == b'&&&':
                    decrypting = receivedata[beginnum:i]
                    original_message1 = p2pkey['private_key'].decrypt(
                        decrypting,
                        padding.OAEP(
                            mgf=padding.MGF1(algorithm=hashes.SHA256()),
                            algorithm=hashes.SHA256(),
                            label=None
                        )
                    )
                    original_message = original_message + original_message1
                    beginnum = i + 3
                i = i + 1
            decrypting = receivedata[beginnum:]
            original_message1 = p2pkey['private_key'].decrypt(
                decrypting,
                padding.OAEP(
                    mgf=padding.MGF1(algorithm=hashes.SHA256()),
                    algorithm=hashes.SHA256(),
                    label=None
                )
            )
            original_message = original_message + original_message1
        else:
            original_message = p2pkey['private_key'].decrypt(
                receivedata,
                padding.OAEP(
                    mgf=padding.MGF1(algorithm=hashes.SHA256()),
                    algorithm=hashes.SHA256(),
                    label=None
                )
            )
    except:
        p2psituation = 0
        p2plabel.config(text='当前状态：服务器中转')
        return ''

    else:
        #print(original_message)
        return str(original_message,encoding='utf-8')
def p2pencryptplus(data,p2pkey):
    global tcpsituation1,p2psituation
    data = bytes(data, encoding='utf-8')
    #print('abccc')
    #print(p2pkey)

    try:
     encrypted = b''
     if len(data) <= 100:
        encrypted = p2pkey['friendkey'].encrypt(
            data,
            padding.OAEP(
                mgf=padding.MGF1(algorithm=hashes.SHA256()),
                algorithm=hashes.SHA256(),
                label=None
            )
        )
     while len(data) > 100:
        data1 = data[:100]
        encrypting = p2pkey['friendkey'].encrypt(
            data1,
            padding.OAEP(
                mgf=padding.MGF1(algorithm=hashes.SHA256()),
                algorithm=hashes.SHA256(),
                label=None
            )
        )
        encrypted = encrypted + b'&&&' + encrypting
        data = data[100:]
        if len(data) <= 100:
            encrypting = p2pkey['friendkey'].encrypt(
                data,
                padding.OAEP(
                    mgf=padding.MGF1(algorithm=hashes.SHA256()),
                    algorithm=hashes.SHA256(),
                    label=None
                )
            )
            encrypted = encrypted + b'&&&' + encrypting
            break
     return encrypted
    except:
        p2psituation = 0
        print('defeat')
        p2plabel.config(text='当前状态：服务器中转')
        return ''


def shezhi(event):
    global shezhi_on
    if shezhi_on==1:
     t1 = threading.Thread(target=shezhiwindow)
     t1.setDaemon(True)
     t1.start()


def shezhiwindow():
    global jinggao
    global window2
    global slabel1
    global slabel2
    global slabel3
    global slabel4
    global slabel5
    global slabel6
    global slabel7,slabel8
    global soff
    global simagemax
    global scodeentry
    global snichengentry
    global sleft
    global sright
    global sbaocun
    global simagemin,shezhi_on
    jinggao=0
    shezhi_on=0

    window2 = tk.Toplevel()
    window2.overrideredirect(True)
    window2.title('设置')
    window2.geometry('300x530+400+200')
    simagemax = tk.Label(window2)
    slabel1 = tk.Label(window2, text='Pansis账户ID')
    slabel2 = tk.Label(window2, text=userid)
    slabel3 = tk.Label(window2, text='账户删除')
    slabel4 = tk.Label(window2, text='密码')
    slabel5 = tk.Label(window2, text='昵称')
    slabel6 = tk.Label(window2, text='主题')
    slabel7 = tk.Label(window2, text='所有聊天记录将全部删除\n再次点击确认删除',fg='red')
    slabel8 = tk.Label(window2, text='不能含有*&=特殊符号', fg='red')
    scodeentry = tk.Entry(window2, width=15, bd=0)
    snichengentry = tk.Entry(window2, width=15, bd=0)
    snichengentry.delete(0, END)
    snichengentry.insert(END,ws["B1"].value)
    scodeentry.delete(0, END)
    # image1=tk.PhotoImage(file="组件/ui/1/展示背景1.png")
    # image2=tk.PhotoImage(file="组件/ui/1/展示背景2.png")
    # image3=tk.PhotoImage(file="组件/ui/2/展示背景3.png")
    # image4=tk.PhotoImage(file="组件/ui/2/展示背景4.png")
    sleft = tk.Label(window2)
    sright = tk.Label(window2)
    soff = tk.Label(window2)
    simagemin = tk.Label(window2)

    sbaocun = tk.Button(window2, bg='#298CF4', relief=FLAT, width=25, command=baocun1, text='保存', fg='white', bd=0,activebackground='#298CF4')
    simagemax.place(x=-2, y=-2)
    slabel1.place(x=5, y=50)
    slabel2.place(x=100, y=50)
    slabel3.place(x=160, y=50)
    slabel4.place(x=5, y=100)
    slabel5.place(x=5, y=150)
    slabel6.place(x=5, y=200)
    soff.place(x=1, y=3)
    sleft.place(x=5, y=240)
    sright.place(x=245, y=240)
    simagemin.place(x=35, y=230)

    sbaocun.place(x=30, y=450)
    scodeentry.place(x=100, y=100)
    snichengentry.place(x=100, y=150)

    # offimage11 = tk.PhotoImage(file="组件/ui/1/关闭1.png")
    # offimage10 = tk.PhotoImage(file="组件/ui/1/关闭0.png")
    # offimage11 = tk.PhotoImage(file="组件/ui/1/关闭1.png")
    soff.bind("<Button-1>", swindowoff1)

    slabel3.bind("<Button-1>", suserdelete)
    soff.bind("<Enter>", swindowoff2)
    soff.bind("<Leave>", swindowoff3)
    sleft.bind("<Button-1>", swindowleft1)
    sleft.bind("<Enter>", swindowleft2)
    sleft.bind("<Leave>", swindowleft3)
    sright.bind("<Button-1>", swindowright1)
    sright.bind("<Enter>", swindowright2)
    sright.bind("<Leave>", swindowright3)
    sui()



def sui():

    global soffimage0
    global soffimage1
    global srightimage0
    global srightimage1
    global sleftimage0
    global sleftimage1
    global simagemaximage
    global simageminimage


    if int(uiid) <= 2:
        uimain = '1'
    else:
        uimain = '2'
  #  print('uiid:'+uiid)
  #  print('uiid:'+uimain)
    simagemaximage = tk.PhotoImage(file="组件/ui/" + uimain + "/注册背景" + uiid + ".png")
    simageminimage = tk.PhotoImage(file="组件/ui/" + uimain + "/展示背景" + uiid + ".png")
 #   print("组件/ui/" + uimain + "/展示背景" + uiid + ".png")
  #  print("组件/ui/" + uimain + "/注册背景" + uiid + ".png")
    soffimage0 = tk.PhotoImage(file="组件/ui/" + uimain + "/返回0.png")
    soffimage1 = tk.PhotoImage(file="组件/ui/" + uimain + "/返回1.png")
    sleftimage0 = tk.PhotoImage(file="组件/ui/" + uimain + "/左翻页0.png")
    sleftimage1 = tk.PhotoImage(file="组件/ui/" + uimain + "/左翻页1.png")
    srightimage0 = tk.PhotoImage(file="组件/ui/" + uimain + "/右翻页0.png")
    srightimage1 = tk.PhotoImage(file="组件/ui/" + uimain + "/右翻页1.png")
    simagemax.config(image=simagemaximage)
    simagemin.config(image=simageminimage)
    sright.config(image=srightimage0)
    sleft.config(image=sleftimage0)
    if uimain == '1':
        soff.config(image=soffimage0, bg='black')
        scodeentry.config(bg='#555555')
        snichengentry.config(bg='#555555')
        slabel1.config(bg='black', fg='#8E8E84')
        slabel2.config(bg='black', fg='#8E8E84')
        slabel3.config(bg='black', fg='#8E8E84')
        slabel4.config(bg='black', fg='#8E8E84')
        slabel5.config(bg='black', fg='#8E8E84')
        slabel6.config(bg='black', fg='#8E8E84')
        slabel8.config(bg='black')
        slabel7.config(bg='black')
        sleft.config(bg='black')
        sright.config(bg='black')
        simagemin.config(bg='black')
        simagemax.config(bg='black')
        window2.config(bg='black')
    if uimain == '2':
            soff.config(image=soffimage0, bg='white')
            scodeentry.config(bg='#F0F0F0')
            snichengentry.config(bg='#F0F0F0')
            slabel1.config(bg='white', fg='#8E8E84')
            slabel2.config(bg='white', fg='#8E8E84')
            slabel3.config(bg='white', fg='#8E8E84')
            slabel4.config(bg='white', fg='#8E8E84')
            slabel5.config(bg='white', fg='#8E8E84')
            slabel6.config(bg='white', fg='#8E8E84')
            slabel8.config(bg='white')
            slabel7.config(bg='white')
            sleft.config(bg='white')
            sright.config(bg='white')
            simagemin.config(bg='white')
            simagemax.config(bg='white')
            window2.config(bg='white')






def swindowoff2(event):
    soff.config(image=soffimage1)


def swindowoff3(event):
    soff.config(image=soffimage0)


def swindowoff1(event):  ###
  #  print('1')
    global shezhi_on
    shezhi_on=1
    window2.destroy()


def swindowright2(event):
    sright.config(image=srightimage1)


def swindowright3(event):
    sright.config(image=srightimage0)


def swindowright1(event):  ###
    global uiid
    suiid = int(uiid)
    if suiid < 4:
        suiid = suiid + 1
        uiid = str(suiid)
        sui()
        userui(uiid)


def swindowleft2(event):
    sleft.config(image=sleftimage1)


def swindowleft3(event):
    sleft.config(image=sleftimage0)


def swindowleft1(event):  ###
    global uiid
    suiid=int(uiid)
    if suiid>1:
        suiid=suiid-1
        uiid=str(suiid)
        sui()
        userui(uiid)


def baocun1():  ###s
    global usernicheng,logcode,shezhi_on
    data ="0017"+userid
 #   print('1')
    if snichengentry.get() != ws['B1'].value:
        usernicheng = snichengentry.get()
        data=data+usernicheng+"*"
        ws['B1'].value= usernicheng
        nicheng.config(text=usernicheng)
    else:data=data+"&*"
    data = data +uiid+"*"
    if scodeentry.get()!='':
        newcode=scodeentry.get()
        i=0
        while (i<=len(newcode)-1):
            if newcode[i]=='&' or newcode[i]=='*' or newcode[i]=='=':
                slabel8.place(x=100, y=180)
                return
            i=i+1
        #print(newcode)
        newcode =hashlib.md5(newcode.encode('utf8')).hexdigest()
        data = data + newcode+'&'
        #print(newcode)
        logcode=newcode
    else:data=data+"&"
    #print(data)
    data=encryptplus(data)
    tcps.send(data)
    shezhi_on = 1
    window2.destroy()


suserdeletenum=0
def suserdelete(event):
    global suserdeletenum
    if suserdeletenum==0:
        suserdeletenum=1
        slabel3.config(text='所有聊天记录将全部删除\n再次点击确认删除',fg='red')
    else:

        window2.destroy()
        try:
            os.remove(userid + ".xlsx")
        except:
            pass
        sys.exit(1)
















def dengluok():          #登录成功事件


    global ws
    global wb

    try:
        wb = load_workbook(userid+".xlsx")
    except:
        wb = Workbook()  # 新建表格()
        ws = wb.active  # 赋值当前活动sheet
        #print(ws.title)  # 打印ws的sheet的标题
        ws.title = "Sheet1"
        ws["A1"].value=userid

    else:
        wb.active = wb["Sheet1"]
        ws = wb.active
    userinput.place_forget()
    codeinput.place_forget()
    denglu.place_forget()
    codeshow.place_forget()
    codeshow.pack_forget()
    denglusituation.place_forget()
    global nicheng
    global uiid
    nicheng=tk.Label(window)
    nicheng.config(text=' ')
    addnicheng(usernicheng)
    uiid = str(uiid)
    ws['C1'].value = uiid
    wb.save(userid + ".xlsx")
    #print('uiid' + uiid)



    global newchatlable
    global newchatentry
    global userexit
    global shezhilabel
    global useridlabel
    window.geometry('800x530+400+200')
    windowmin.place(x=725,y=2)
    windowoff.place(x=755,y=2)
    nicheng.place(x=10, y=40)

    tcpsituation.place(x=110, y=40)
    newchatlable=tk.Label(window,text='发起新聊天')
    newchatentry=tk.Entry(window,text='请输入对方账号',bd=0,width=23)
    useridlabel=tk.Label(window,text=userid)
    shezhilabel = tk.Label(window, text='设置')
    userexit=tk.Label(window,text='注销')
    newchatentry.insert(END,'请输入对方账号')
    newchatentry.bind('<FocusIn>',newchatentryfocusin)
    newchatentry.bind('<FocusOut>',newchatentryfocusout)
    newchatentry.bind("<Return>",newchat)
    userexit.bind("<Button-1>",windowoff1)
    shezhilabel.bind("<Button-1>",shezhi)
    newchatlable.place(x=10,y=100)
    newchatentry.place(x=90, y=102)
    userexit.place(x=70, y=70)
    useridlabel.place(x=10, y=70)
    shezhilabel.place(x=110,y=70)
    friendline1()
    userui(uiid)
    friendlistsituation(0)

def userui(uiid):
    #print('uiid为'+uiid)
    if int(uiid)<=2:
        uimain='1'
    else:uimain='2'
    global indexbgimage
    global offimage0
    global minimage0
    global offimage1
    global minimage1
    global useridlabel
    global sendenter,p2penter
    indexbgimage = tk.PhotoImage(file="组件/ui/"+uimain+"/用户背景"+uiid+".png")
    offimage0 = tk.PhotoImage(file="组件/ui/" + uimain + "/关闭0.png")
    minimage0 = tk.PhotoImage(file="组件/ui/" + uimain + "/最小化0.png")
    offimage1 = tk.PhotoImage(file="组件/ui/" + uimain + "/关闭1.png")
    minimage1 = tk.PhotoImage(file="组件/ui/" + uimain + "/最小化1.png")
    indexbgi.config(image=indexbgimage)
    if uimain=='1':
        indexbgi.config(bg='black')
        windowoff.config(image=offimage0,bg='black')
        windowmin.config(image=minimage0,bg='black')
        userinput.config(bg='#555555')
        codeinput.config(bg='#555555')
        tcpsituation.config(fg='#8E8E84',bg='black')
        result.config(bg='#2B2B2B',fg='#FFFFFF')
        setput.config(bg='#535353',fg='#FFFFFF')
        p2penter.config(bg='#535353', fg='#FFFFFF')
        p2plabel.config(bg='#535353', fg='#FFFFFF')
        sendenter .config(bg='#535353', fg='#FFFFFF')
        codeshow.config(bg='black',fg='#8E8E84')
        nicheng.config(bg='black',fg='#8E8E84')
        newchatlable.config(bg='black', fg='#8E8E84')
        useridlabel.config(bg='black', fg='#8E8E84')
        shezhilabel.config(bg='black', fg='#8E8E84')
        userexit.config(bg='black',fg='#8E8E84')
        newchatentry.config(bg='#535353',fg='#8E8E84')
        friendlist1.config(bg='#2B2B2B',fg='#8E8E84')
        friendlist2.config(bg='#2B2B2B', fg='#8E8E84')
        friendlist3.config(bg='#2B2B2B', fg='#8E8E84')
        friendlist4.config(bg='#2B2B2B', fg='#8E8E84')
        friendlist5.config(bg='#2B2B2B', fg='#8E8E84')
        friendlist6.config(bg='#2B2B2B', fg='#8E8E84')
        friendlist7.config(bg='#2B2B2B', fg='#8E8E84')
        friendlist8.config(bg='#2B2B2B', fg='#8E8E84')
        window.config(bg='black')
    if uimain=='2':
        indexbgi.config(bg='white')
        windowoff.config(image=offimage0,bg='white')
        windowmin.config(image=minimage0,bg='white')
        userinput.config(bg='#F0F0F0')
        codeinput.config(bg='#F0F0F0')
        tcpsituation.config(fg='#8E8E84',bg='white')
        result.config(bg='#C1C1C1',fg='#747B74')
        setput.config(bg='#EBECDE',fg='#747B74')
        sendenter.config(bg='#EBECDE', fg='#747B74')
        p2penter.config(bg='#EBECDE', fg='#747B74')
        p2plabel.config(bg='#EBECDE', fg='#747B74')
        codeshow.config(bg='white',fg='#8E8E84')
        nicheng.config(bg='white',fg='#8E8E84')
        newchatlable.config(bg='white', fg='#8E8E84')
        useridlabel.config(bg='white', fg='#8E8E84')
        shezhilabel.config(bg='white', fg='#8E8E84')
        userexit.config(bg='white',fg='#8E8E84')
        newchatentry.config(bg='#EBECDE',fg='#8E8E84')
        friendlist1.config(bg='#EBECDE',fg='#8E8E84')
        friendlist2.config(bg='#EBECDE', fg='#8E8E84')
        friendlist3.config(bg='#EBECDE', fg='#8E8E84')
        friendlist4.config(bg='#EBECDE', fg='#8E8E84')
        friendlist5.config(bg='#EBECDE', fg='#8E8E84')
        friendlist6.config(bg='#EBECDE', fg='#8E8E84')
        friendlist7.config(bg='#EBECDE', fg='#8E8E84')
        friendlist8.config(bg='#EBECDE', fg='#8E8E84')
        window.config(bg='white')













def userexit1(event):
    global indexbgimage
    global offimage0
    global minimage0
    global offimage1
    global minimage1
    global useridlabel
    userui('1')
    codeshow.deselect()
    window.geometry('510x258+400+200')
    indexbgimage = tk.PhotoImage(file="组件/ui/1/登录背景1.png")
    indexbgi.config(image=indexbgimage)


    userinput.place(x=300, y=50)
    codeinput.place(x=300, y=85)
    denglu.place(x=300, y=160)
    codeshow.place(x=400, y=113)

    tcpsituation.place(x=310, y=115)
    windowoff.place(x=465, y=2)
    windowmin.place(x=435, y=2)
    nicheng.place_forget()

    useridlabel.place_forget()
    shezhilabel.place_forget()
    newchatlable.place_forget()
    newchatentry.place_forget()
    friendlist1.place_forget()
    friendlist2.place_forget()
    friendlist3.place_forget()
    friendlist4.place_forget()
    friendlist5.place_forget()
    friendlist6.place_forget()
    friendlist7.place_forget()
    friendlist8.place_forget()
    result.place_forget()
    setput.place_forget()
    sendenter.place_forget()
    p2penter.place_forget()
    p2plabel.place_forget()
    userexit.place_forget()
    newchatentry.delete(0, END)
    global userid
    global nowfriendid
    nowfriendid=''
    userid=''



def newchatentryfocusin(event):
    global newchatlable
    global newchatentry
    if newchatentry.get()=='请输入对方账号' or newchatentry.get()=='请求已发送，等待对方回信' or newchatentry.get()=='请输入正确的账号':
        newchatentry.delete(0, END)
def newchatentryfocusout(event):
    global newchatlable
    global newchatentry
    if newchatentry.get() == '':
        newchatentry.insert(END,'请输入对方账号')



def newchat(event):
    global newchatlable
    global newchatentry
    global newchatlablesendend
    data=time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())+'&你好'
    data = encrypt1(data, code)

    if newchatentry.get().isnumeric() :
        try:
            tcps.send(('0011' + str(newchatentry.get()) + '&' + userid + '&' + str(data)).encode('utf-8'))
        except:
            a = 1
        # print('0011'+str(newchatentry.get())+userid+'&'+time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())+'&'+'你好')
        newchatentry.delete(0, END)
        newchatentry.insert(END, '请求已发送，等待对方回信')
    else:
        newchatentry.delete(0, END)
        newchatentry.insert(END, '请输入正确的账号')



def friendline1():
    global friendlist1
    global friendlist2
    global friendlist3
    global friendlist4
    global friendlist5
    global friendlist6
    global friendlist7
    global friendlist8
    global firstlist
    global friendlistmainy
    global friendlistmainymin
    global friendlistmainsum
    global friendlistmainx
    global friendlistmain
    global friendline
    global friendlinetime
    friendline=[]
    friendlinetime = []
    friendlinetime1 = []
    for col in ws.iter_cols(min_row=2, min_col=1, max_row=999, max_col=1):
        for cell in col:
            if cell.value!=None:
                friendlinetime1.append(ws["C"+str(cell.row)].value)
            else:break
    friendlinetime=sorted(friendlinetime1,reverse=True)
    i=0
    while i<len(friendlinetime):
        for col in ws.iter_cols(min_row=2, min_col=3, max_row=999, max_col=3):
            for cell in col:
                if cell.value == friendlinetime[i]:

                    friendline.append(ws["A" + str(cell.row)].value)
                if cell.value==None:
                    break
        i=i+1
    i=0
  #  print(friendline)
  #  print(friendlinetime)
    friendlist1 = tk.Label(window, height=2, bg='blue', width=40, text='1')
    friendlist2 = tk.Label(window, height=2, bg='blue', width=40, text='2')
    friendlist3 = tk.Label(window, height=2, bg='blue', width=40, text='3')
    friendlist4 = tk.Label(window, height=2, bg='blue', width=40, text='4')
    friendlist5 = tk.Label(window, height=2, bg='blue', width=40, text='5')
    friendlist6 = tk.Label(window, height=2, bg='blue', width=40, text='6')
    friendlist7 = tk.Label(window, height=2, bg='blue', width=40, text='7')
    friendlist8 = tk.Label(window, height=2, bg='blue', width=40, text='8')


    friendlist1.bind(" <Button-1>",friendlist1click)
    friendlist2.bind(" <Button-1>", friendlist2click)
    friendlist3.bind(" <Button-1>", friendlist3click)
    friendlist4.bind(" <Button-1>", friendlist4click)
    friendlist5.bind(" <Button-1>", friendlist5click)
    friendlist6.bind(" <Button-1>", friendlist6click)
    friendlist7.bind(" <Button-1>", friendlist7click)
    friendlist8.bind(" <Button-1>", friendlist8click)

    firstlist = friendlist1

    friendlistmainymin = 170
    friendlistmain = 1
    friendlistmainy = friendlistmainymin
    friendlistmainx=15

    if len(friendline)>=9:
        global friendlist8id
        global friendlist1id
        global friendlist2id
        global friendlist3id
        global friendlist4id
        global friendlist5id
        global friendlist6id
        global friendlist7id
        friendlist1id = 1
        friendlist2id = 2
        friendlist3id = 3
        friendlist4id = 4
        friendlist5id = 5
        friendlist6id = 6
        friendlist7id = 7
        friendlist8id = 8
        friendlistmainsum = 1
        friendlist1.config(text=str(friendline[friendlist1id - 1]))
        friendlist2.config(text=str(friendline[friendlist2id - 1]))
        friendlist3.config(text=str(friendline[friendlist3id - 1]))
        friendlist4.config(text=str(friendline[friendlist4id - 1]))
        friendlist5.config(text=str(friendline[friendlist5id - 1]))
        friendlist6.config(text=str(friendline[friendlist6id - 1]))
        friendlist7.config(text=str(friendline[friendlist7id - 1]))
        friendlist8.config(text=str(friendline[friendlist8id - 1]))
        friendlistmove()
        window.bind('<MouseWheel>', mousemove)
    else:
        friendlistcalm()

def friendlistcalm():
    global friendlist1
    global friendlist2
    global friendlist3
    global friendlist4
    global friendlist5
    global friendlist6
    global friendlist7
    global friendlist8
    global firstlist
    global friendlistmainy
    global friendlistmain
    global friendline
    global friendlistmainx
    global friendlinetime
    global friendlistmain
    global friendlistmainsum
    global friendlist8id
    global friendlist1id
    global friendlist2id
    global friendlist3id
    global friendlist4id
    global friendlist5id
    global friendlist6id
    global friendlist7id
    friendlist1id=1
    friendlist2id = 2
    friendlist3id = 3
    friendlist4id = 4
    friendlist5id = 5
    friendlist6id = 6
    friendlist7id = 7
    friendlist8id = 8
    if len(friendline)>=1:
        #friendlist1 = tk.Label(window, height=2, bg='blue', width=40, text='1')
        friendlist1.config(text=str(friendline[friendlist1id-1]))
        friendlist1.place(x=friendlistmainx, y=friendlistmainy)
        if len(friendline) >= 2:
            #friendlist2 = tk.Label(window, height=2, bg='blue', width=40, text='2')
            friendlist2.config(text=str(friendline[friendlist2id-1]))
            friendlist2.place(x=friendlistmainx, y=friendlistmainy+45*1)
            if len(friendline) >= 3:
                #friendlist3 = tk.Label(window, height=2, bg='blue', width=40, text='3')
                friendlist3.config(text=str(friendline[friendlist3id-1]))
                friendlist3.place(x=friendlistmainx, y=friendlistmainy+45*2)
                if len(friendline) >= 4:
                    #friendlist4 = tk.Label(window, height=2, bg='blue', width=40, text='4')
                    friendlist4.config(text=str(friendline[friendlist4id-1]))
                    friendlist4.place(x=friendlistmainx, y=friendlistmainy+45*3)
                    if len(friendline) >= 5:
                        #friendlist5 = tk.Label(window, height=2, bg='blue', width=40, text='5')
                        friendlist5.config(text=str(friendline[friendlist5id-1]))
                        friendlist5.place(x=friendlistmainx, y=friendlistmainy+45*4)
                        if len(friendline) >= 6:
                            #friendlist6 = tk.Label(window, height=2, bg='blue', width=40, text='6')
                            friendlist6.config(text=str(friendline[friendlist6id-1]))
                            friendlist6.place(x=friendlistmainx, y=friendlistmainy+45*5)
                            if len(friendline) >= 7:
                                #friendlist7 = tk.Label(window, height=2, bg='blue', width=40, text='7')
                                friendlist7.config(text=str(friendline[friendlist7id-1]))
                                friendlist7.place(x=friendlistmainx, y=friendlistmainy+45*6)
                                if len(friendline) >= 8:
                                    #friendlist8 = tk.Label(window, height=2, bg='blue', width=40, text='8')
                                    friendlist8.config(text=str(friendline[friendlist8id-1]))
                                    friendlist8.place(x=friendlistmainx, y=friendlistmainy+45*1)


    
    













def addnicheng(nicheng1):    #为新用户添加昵称 #修改本地昵称
    global usernicheng
    print(nicheng1)
    ws["B1"].value=nicheng1
    usernicheng=nicheng1
    nicheng.config(text=ws["B1"].value)
    wb.save(userid + ".xlsx")



def nouser(): ##登录无此用户事件
    denglusituation.config(text='账号或密码错误')
    denglusituation.place(x=310, y=135)
def codeerror(): ##登录密码错误事件
    denglusituation.config(text='账号或密码错误')
    denglusituation.place(x=310, y=135)

def tcp(senddata):  #数据发送与监听事件


    #print(tcpsituation1)

    #print(senddata)
    senddata=encryptplus(senddata)

    try:
           tcps.send(senddata)

    except:
           a = 1
    else:
        print('发送成功')

def ser0024(receivedata):
    global server_public_key,server_public_key1,server_public_key2,tcpsituation1,friendlistclick_on
    server_public_key = receivedata[4:]
    server_public_key1 = serialization.load_pem_public_key(server_public_key, backend=default_backend())
    #print(server_public_key1)
    #print(server_public_key)
    tcpsituation.config(text='连接服务器成功')
    tcpsituation1 = 1
    friendlistclick_on=1
    if logcode!='':
        tcp('0025'+userid+'&'+logcode)
def decryptplus(receivedata):
    global tcpsituation1
    #print(b'abcdef'+receivedata)

    try:
        original_message = b''
        if receivedata[:3] == b'&&&':
            i = 3
            beginnum = 3
            while i < len(receivedata):


                if receivedata[i:i + 3] == b'&&&':
                    decrypting = receivedata[beginnum:i]
                    original_message1 = client_private_key2.decrypt(
                        decrypting,
                        padding.OAEP(
                            mgf=padding.MGF1(algorithm=hashes.SHA256()),
                            algorithm=hashes.SHA256(),
                            label=None
                        )
                    )
                    original_message = original_message + original_message1
                    beginnum = i + 3
                i = i + 1
            decrypting = receivedata[beginnum:]
            original_message1 = client_private_key2.decrypt(
                decrypting,
                padding.OAEP(
                    mgf=padding.MGF1(algorithm=hashes.SHA256()),
                    algorithm=hashes.SHA256(),
                    label=None
                )
            )
            original_message = original_message + original_message1
        else:
            original_message = client_private_key2.decrypt(
                receivedata,
                padding.OAEP(
                    mgf=padding.MGF1(algorithm=hashes.SHA256()),
                    algorithm=hashes.SHA256(),
                    label=None
                )
            )
    except:
        tcpsituation.config(text='连接服务器失败')
        tcpsituation1 = 0
        return ''

    else:
        #print(original_message)
        return str(original_message,encoding='utf-8')
def encryptplus(data):
    global tcpsituation1
    data = bytes(data, encoding='utf-8')
    try:
     encrypted = b''
     if len(data) <= 100:
        encrypted = server_public_key1.encrypt(
            data,
            padding.OAEP(
                mgf=padding.MGF1(algorithm=hashes.SHA256()),
                algorithm=hashes.SHA256(),
                label=None
            )
        )
     while len(data) > 100:
        data1 = data[:100]
        encrypting = server_public_key1.encrypt(
            data1,
            padding.OAEP(
                mgf=padding.MGF1(algorithm=hashes.SHA256()),
                algorithm=hashes.SHA256(),
                label=None
            )
        )
        encrypted = encrypted + b'&&&' + encrypting
        data = data[100:]
        if len(data) <= 100:
            encrypting = server_public_key1.encrypt(
                data,
                padding.OAEP(
                    mgf=padding.MGF1(algorithm=hashes.SHA256()),
                    algorithm=hashes.SHA256(),
                    label=None
                )
            )
            encrypted = encrypted + b'&&&' + encrypting
            break
     return encrypted
    except:
        tcpsituation.config(text='连接服务器失败')
        tcpsituation1 = 0
        return ''
def tcpreceive():
 global datatime
 while True:
  global code
  global userid
  global m
  global uiid
  global usernicheng
  global filelistnonum,logcode,logcode1,gip,gport,p2pownkey
  if (update.nowsituation==0):
    try:
     receivedata = tcps.recv(10240)  # 等待接受(数据最大为1024)

    except:
        continue
    else:
     if receivedata[0:4] == bytes('0024', encoding='utf-8'):
            ser0024(receivedata)
            continue
     if not (receivedata[0:4].isdigit()):
         receivedata=decryptplus(receivedata)
     else:
         receivedata = receivedata.decode('utf-8')
     指令 = receivedata[0:4]
     print(指令)
     if 指令=='0002':
         i = 30
         typenum = 1
         data=receivedata
         massage1=''
         while i <= len(data) - 1:
             if data[i] == "*":
                 if typenum == 1:
                     code = data[4:i]
                     #print('1234'+code)
                     #code=decrypt1(code,logcode)
                     #print('1234' + code)
                 if typenum == 2:
                     usernicheng = data[end + 1:i]
                 if typenum == 3:
                     uiid = data[end + 1:i]
                     massage1 = data[i + 1:]
                     break
                 end = i
                 typenum = typenum + 1
             i = i + 1
         userid=userinput.get()
         logcode=logcode1
         dengluok()
         massage(massage1)

     if 指令=='0003':
           nouser()
           logcode = ''
     if 指令=='0004':                                                                   #未进行
           codeerror()
           logcode = ''
    # if 指令=='0010':
    #       print(receivedata[4:])
     #      addnicheng(receivedata[4:])
     if 指令=='0014':
           global friendnicheng
           friendnicheng=receivedata[4:]
           friendlistclick2()

     if 指令 == '0006':
           #print("12345")
           massage(receivedata)
     if 指令 == '0021':

         try:
             abc=fileclient.uploadport
         except:
             abc=''
         else:
          if receivedata[4:8]==fileclient.uploadport:
            ser0021(receivedata[8:])

     if (指令=='0027'):
         p2pbegin(receivedata[4:])
     if (指令=='0031'):
         #print('743289' + receivedata[4:])
         if updatesituation != 1:
            update1(receivedata[4:])
  else:
      break
 return
updatesituation=0
def update1(url):
  global updatesituation
  if updatesituation!=1:
    updatesituation = 1

    print(url)
    print(sys.path[0]+'\\组件\\update.exe')
    f = open('组件/url.txt', 'w')
    f.write(url)
    f.close()
    a=0
    window.withdraw()
    update.main()
    appexit1 = threading.Thread(target=appexit)
    appexit1.setDaemon(True)
    appexit1.start()

def appexit():     #用于强制退出主程序的函数调用，务必用多线程
    while (update.nowsituation!=2):
        time.sleep(1)
    window.destroy()





def ser0021(receivedata):
           global filelistnonum
           global newuploadfileline,friendlistclick_on
           datatime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
           filename=receivedata
           result.config(state='normal')
           result.delete(str(newuploadfileline)+'.0',str(newuploadfileline)+'.end')
           buttontext = filename + "(已上传)  打开  打开文件夹  删除文件"
           result.insert(str(newuploadfileline)+'.0', buttontext + '\n')
           #  resultlinenum = resultlinenum + 2
           filelistline[filename] = newuploadfileline  # ！！！！！！！！！！！！！！1切换删除
           filelistno[filename] = filelistnonum  # ！！！！！！！！！！！！！！1切换删除
           filelistnonum=filelistnonum+1
           friendlistclick_on = 1

           if filename[0] == '(':
               i = 1
               while i < len(filename):
                   if filename[i] == ')':
                       filename1 = filename[i + 1:]
                       filelocation[filename]=filelocation[filename1]

                       break
                   i = i + 1
           downloadedaddfiletag(filename)
           data = '$$$' + filename + '&' + filelocation[filename]
           data1 = '$$$' + filename + '&*'
           for col in ws.iter_cols(min_row=2, min_col=1, max_row=9999, max_col=1):
               for cell in col:
                   if cell.value == nowfriendid:
                       cellrow = cell.row
                       break
           for row in ws.iter_rows(min_row=cellrow, min_col=11, max_row=cellrow, max_col=9999):  # 括号代表遍历第一行到第二行,第二列到第三列
               for cell in row:
                   if cell.value == None:
                       data = datatime + '&' + userid + data
                       #print('而' + data)
                       data = encrypt1(data, code)
                       #print('而我却' + data)
                       cell.value = data
                       wb.save(userid + '.xlsx')
                       break
           result.config(state='disabled')
           data1=datatime + '&' + data1
           data1 = encrypt1(data1, code)
           data1 = '0011' + nowfriendid + '&' + userid + '&' +  data1
           tcp(data1)
           result.see(tk.END)


def massage(receivedata):
    # print("12345")

    m = 0.001
    receivedata1 = receivedata
    # print(receivedata1)
    i = 4
    list = ['123']
    databegin = 4
    global dataend,filelistnonum
    global nowfriendid
    dataend = 0
    while i <= len(receivedata1) - 1:
        if receivedata1[i] == '&':
            dataend = i

            list.append(receivedata1[databegin:dataend])

            databegin = i + 1
        i = i + 1
    list.append(receivedata1[databegin:])
    i = 1
    mm = 0
    sum1 = 0
    datatime=''
    datatimebegin=0

    if list[1] != '':
        while i <= len(list) - 1:

            data = decrypt1(list[i], code)

            #print("12345a"+data)
            friendid = data[0:5]

            while mm < len(data):
                if data[mm] == '&':
                    if sum1 == 0:
                        sum1 = 1
                        datatimebegin = mm

                    elif sum1 == 1:
                        datatime = data[datatimebegin + 1:mm]
                        data = data[mm + 1:]
                        sum1 = 0
                        break
                mm = mm + 1
            mm = 0
            sum1 = 0
            data = datatime + '&' + friendid + data
            data = encrypt1(data, code)
            for col in ws.iter_cols(min_row=2, min_col=1, max_row=999, max_col=1):
                for cell in col:
                    cellrow = cell.row
                    # print(friendid)
                    # print(cell.value)
                    if cell.value == friendid:

                        #print('检索到用户')
                        for row in ws.iter_rows(min_row=cellrow, min_col=11, max_row=cellrow,
                                                max_col=999):  # 括号代表遍历第一行到第二行,第二列到第三列
                            for cell in row:
                                if cell.value == None:
                                    cell.value = data
                                    ws["C" + str(cell.row)].value = time.time()
                                    ws["C" + str(cell.row)].value = ws["C" + str(cell.row)].value - m
                                    if nowfriendid != friendid:
                                        ws["D" + str(cell.row)].value='1'
                                    if nowfriendid == friendid:
                                        data = decrypt1(cell.value, code)
                                        # print("!!!"+data)
                                        i = 0

                                        while i <= len(data) - 1:
                                            if data[i] == '&':
                                                dataid = data[i + 1:i + 6]
                                                datatime = data[:i]
                                                data = data[i + 6:]
                                                #print(data)

                                                if dataid == userid:
                                                    dataid = usernicheng + '(我)'

                                                else:
                                                    dataid = dataid + '  ' + str(friendnicheng) + '(对方)'
                                                if data[:3] == '$$$':
                                                    a = 2
                                                    while a <= len(data) - 1:
                                                        if data[a] == '&':
                                                            result.config(state='normal')

                                                            filelocation[data[3:a]] = data[
                                                                                      a + 1:]  # data[0:a]文件名  data[a+1:]文件路径     #！！！！！！！！！！！！！！1切换删除
                                                            #print("uuds\n")
                                                            #print(filelocation[data[3:a]])
                                                            if data[a + 1:] == "*":
                                                                buttontext = data[3:a] + "(未下载)  点击下载"
                                                                result.insert('end', dataid + '   ' + datatime + '\n')
                                                                result.insert('end', buttontext + '\n')
                                                                #  resultlinenum = resultlinenum + 2
                                                                #print(result.index('end-1c').split('.')[0])

                                                                filelistline[data[3:a]] = int(
                                                                    result.index('end-1c').split('.')[
                                                                        0]) - 1  # ！！！！！！！！！！！！！！1切换删除
                                                                filelistno[
                                                                    data[3:a]] = filelistnonum  # ！！！！！！！！！！！！！！1切换删除
                                                                downloadaddfiletag(data[3:a], filelistline[data[3:a]])
                                                            else:
                                                                buttontext = data[3:a] + "(已下载)  打开  打开文件夹  删除文件"
                                                                result.insert('end', dataid + '   ' + datatime + '\n')
                                                                result.insert('end', buttontext + '\n')
                                                                #  resultlinenum = resultlinenum + 2
                                                                filelistline[data[3:a]] = int(
                                                                    result.index('end-1c').split('.')[
                                                                        0]) - 1  # ！！！！！！！！！！！！！！1切换删除
                                                                filelistno[
                                                                    data[3:a]] = filelistnonum  # ！！！！！！！！！！！！！！1切换删除
                                                                downloadedaddfiletag(data[3:a])
                                                            break
                                                        a = a + 1
                                                else:
                                                    result.config(state='normal')
                                                    result.insert('end', dataid + '   ' + datatime + '\n' + data + '\n')
                                                # resultlinenum = resultlinenum + 2
                                                break
                                            i = i + 1
                                        filelistnonum = filelistnonum + 1
                                        result.see(tk.END)
                               #     if nowfriendid != '':
                              #          friendlistclick()
                                    #print(time.time() - m)
                                    #print(m)

                                    i = i + 1
                                    m = m + 0.001

                                    break
                            break
                        break
                    elif cell.value == None:
                        #print('未检索到用户')
                        cell.value = friendid
                        ws["K" + str(cell.row)].value = data
                        ws["C" + str(cell.row)].value = time.time()
                        ws["C" + str(cell.row)].value = ws["C" + str(cell.row)].value - m
                        friendline1()
                        userui(uiid)


                        #if nowfriendid != '':
                        #    friendlistclick()

                        # print(m)
                        m = m + 0.001
                        break
        friendlistsituation(0)
        wb.save(userid + ".xlsx")
        m = 0.001

    #print("nh" + str(len(list)))















def denglu1():   #
    global logcode1
    user=userinput.get()
    logcode1=codeinput.get()
    logcode1= hashlib.md5(logcode1.encode('utf8')).hexdigest()

  #  code=encrypt1(code,logcode)
    if user!='' and user!='Pansis账户' and user.isnumeric() :
       a="0001"+userinput.get()+'&'+logcode1
       tcp(a)
    else:
        denglusituation.config(text='请输入账号')
        denglusituation.place(x=310,y=135)

    return
def userinputdelte(event):
    if userinput.get()=='Pansis账户':
        userinput.delete(0, END)
    return
def userinputadd(event):
    if userinput.get()=='':
        userinput.insert(END,'Pansis账户')
    return
def codeinputdelte(event):
    if codeinput.get()=='密码':
        codeinput.delete(0, END)
        codeshow1()
    return
def codeinputadd(event):
    if codeinput.get()=='':
        codeinput.insert(END,'密码')
        codeinput.config(show='')
    return


def windowoff1(event):  ###
  #  print('1')
    sys.exit(1)
    #window.destroy()


def windowmin1(event):    ###
    #print('2')
    window.overrideredirect(False)
    window.iconify()
#def windowshow(event):
#    time.sleep(0.1)
#
#    window.overrideredirect(True)
 #   time.sleep(0.1)
#def windowout(event):
 #   window.overrideredirect(False
def windowoff2(event):
    windowoff.config(image=offimage1)
def windowmin2(event):
    windowmin.config(image=minimage1)
def windowoff3(event):
    windowoff.config(image=offimage0)
def windowmin3(event):
    windowmin.config(image=minimage0)

def windowmove(event):
    new_x = (event.x - window.x) + window.winfo_x()
    new_y = (event.y - window.y) + window.winfo_y()
    s = f"{window.winfo_width()}x{window.winfo_height()}+{new_x}+{new_y}"
    window.geometry(s)
def windowclick(event):
        """获取当前窗口位置并保存"""
        window.x, window.y = event.x, event.y







denglu=tk.Button(window,bg='#298CF4',relief=FLAT,width=25,command=denglu1,text='登录',fg='white',bd=0,activebackground='#298CF4')
codeshow=tk.Checkbutton(window,text='显示密码',command=codeshow1,variable=codeshowvariable,bg='black',fg='#8E8E84', highlightcolor='black')
denglu.place(x=300,y=160)
codeshow.deselect()
codeshow.place(x=400,y=113)
userinput.bind("<FocusOut>",userinputadd)
userinput.bind("<FocusIn>",userinputdelte)
userinput.insert(END,'Pansis账户')
codeinput.bind("<FocusOut>",codeinputadd)
codeinput.bind("<FocusIn>",codeinputdelte)
#setput.bind("<Return>",liaotiansend)
codeinput.insert(END,'密码')
windowoff.bind("<Button-1>",windowoff1)
windowmin.bind("<Button-1>",windowmin1)

windowoff.bind("<Enter>",windowoff2)
windowoff.bind("<Leave>",windowoff3)
windowmin.bind("<Enter>",windowmin2)
windowmin.bind("<Leave>",windowmin3)
indexbgi.bind("<B1-Motion>",windowmove)
indexbgi.bind("<Button-1>",windowclick)
windnd.hook_dropfiles(setput, func=upload)
tcptest = threading.Thread(target=tcpsitiontest)
tcptest.setDaemon(True)
tcpreceivet4=threading.Thread(target=tcpreceive)
tcpreceivet4.setDaemon(True)
tcptest.start()
tcpreceivet4.start()

window.mainloop()
tcps.close                       #关闭套接字